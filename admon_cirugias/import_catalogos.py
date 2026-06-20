"""Plantilla y carga masiva de Doctores y Hospitales (.xlsx)."""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import Doctor, Hospital


COLS_DOCTOR = [
    ('Nombre completo', 'nombre', True),
    ('RFC', 'rfc', False),
    ('Razón social', 'razon_social', False),
    ('Email', 'email', False),
    ('Teléfono', 'telefono', False),
    ('Celular', 'celular', False),
    ('Dirección', 'direccion', False),
    ('Cédula', 'cedula', False),
    ('Activo (Sí/No)', 'activo', False),
]
EJEMPLO_DOCTOR = {
    'nombre': 'Dr. Jorge De La Vega Valdés', 'rfc': 'VEVJ700101AAA',
    'razon_social': 'Jorge De La Vega Valdés', 'email': 'dr.delavega@mail.com',
    'telefono': '6671112233', 'celular': '6679998877', 'direccion': 'Culiacán, Sinaloa',
    'cedula': '1234567', 'activo': 'Sí',
}

COLS_HOSPITAL = [
    ('Código', 'codigo', False),
    ('Nombre', 'nombre', True),
    ('Ciudad', 'ciudad', False),
    ('Activo (Sí/No)', 'activo', False),
]
EJEMPLO_HOSPITAL = {'codigo': 'HGC', 'nombre': 'Hospital General de Culiacán', 'ciudad': 'Culiacán', 'activo': 'Sí'}


def _norm_bool(v):
    if v is None or str(v).strip() == '':
        return True
    return str(v).strip().lower() in ('sí', 'si', 's', 'x', '1', 'true', 'verdadero', 'activo')


def _build(cols, ejemplo, sheet_name, con_ejemplo):
    wb = Workbook(); ws = wb.active; ws.title = sheet_name
    FONT = 'Arial'
    hdr = PatternFill('solid', start_color='1F3864'); hdr_req = PatternFill('solid', start_color='C55A11')
    thin = Side(style='thin', color='D0D0D0'); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for j, (label, key, req) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=j, value=label)
        c.font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
        c.fill = hdr_req if req else hdr
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        ws.column_dimensions[c.column_letter].width = 30 if key in ('nombre', 'direccion', 'razon_social') else 16
    if con_ejemplo:
        for j, (label, key, req) in enumerate(cols, start=1):
            c = ws.cell(row=2, column=j, value=ejemplo.get(key, ''))
            c.font = Font(name=FONT, size=10, italic=True, color='808080')
    ws.freeze_panes = 'A2'; ws.row_dimensions[1].height = 28
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def generar_plantilla_doctores(con_ejemplo=True):
    return _build(COLS_DOCTOR, EJEMPLO_DOCTOR, 'Doctores', con_ejemplo)


def generar_plantilla_hospitales(con_ejemplo=True):
    return _build(COLS_HOSPITAL, EJEMPLO_HOSPITAL, 'Hospitales', con_ejemplo)


def _leer(archivo, cols, sheet_name):
    wb = load_workbook(archivo, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    headers = {}
    for j, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = j
    label_to_key = {label.lower(): key for label, key, _ in cols}
    col = {label_to_key[k]: j for k, j in headers.items() if k in label_to_key}
    return ws, col


def importar_doctores(archivo, empresa):
    ws, col = _leer(archivo, COLS_DOCTOR, 'Doctores')
    if 'nombre' not in col:
        raise ValueError("La plantilla no tiene la columna 'Nombre completo'.")
    creados = actualizados = 0; errores = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def val(k):
            j = col.get(k); v = row[j - 1] if j and j - 1 < len(row) else None
            s = str(v).strip() if v is not None else ''
            return '' if s.upper() == 'NA' else s
        nombre = val('nombre')
        if not nombre:
            continue
        datos = {k: val(k) for k in ('rfc', 'razon_social', 'email', 'telefono', 'celular', 'direccion', 'cedula')}
        datos['email'] = datos['email'] or None
        datos['activo'] = _norm_bool(val('activo'))
        try:
            ex = Doctor.objects.filter(empresa=empresa, nombre=nombre).first()
            if ex:
                for k, v in datos.items():
                    setattr(ex, k, v)
                ex.save(); actualizados += 1
            else:
                Doctor.objects.create(empresa=empresa, nombre=nombre, **datos); creados += 1
        except Exception as e:
            errores.append(f"Fila {i} ({nombre}): {e}")
    return {'creados': creados, 'actualizados': actualizados, 'errores': errores}


def importar_hospitales(archivo, empresa):
    ws, col = _leer(archivo, COLS_HOSPITAL, 'Hospitales')
    if 'nombre' not in col:
        raise ValueError("La plantilla no tiene la columna 'Nombre'.")
    creados = actualizados = 0; errores = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def val(k):
            j = col.get(k); v = row[j - 1] if j and j - 1 < len(row) else None
            s = str(v).strip() if v is not None else ''
            return '' if s.upper() == 'NA' else s
        nombre = val('nombre')
        if not nombre:
            continue
        datos = {'codigo': val('codigo'), 'ciudad': val('ciudad'), 'activo': _norm_bool(val('activo'))}
        try:
            ex = (Hospital.objects.filter(empresa=empresa, codigo=val('codigo')).first()
                  if val('codigo') else None) or Hospital.objects.filter(empresa=empresa, nombre=nombre).first()
            if ex:
                for k, v in datos.items():
                    setattr(ex, k, v)
                ex.save(); actualizados += 1
            else:
                Hospital.objects.create(empresa=empresa, nombre=nombre, **datos); creados += 1
        except Exception as e:
            errores.append(f"Fila {i} ({nombre}): {e}")
    return {'creados': creados, 'actualizados': actualizados, 'errores': errores}
