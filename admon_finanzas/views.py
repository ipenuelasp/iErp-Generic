import decimal

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.views import View
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import Sum

from admon_empresas.models import Moneda
from admon_compras.models import OrdenCompra, Proveedor
from .models import FacturaProveedor, FacturaCliente, CfdiCliente, Pago, AplicacionPago, MetodoPago
from . import services


def _leer_cfdi(xml_file):
    """Extrae subtotal, IVA trasladado, retenciones, total, UUID y RFC receptor
    de un XML CFDI (4.0/3.3). Devuelve dict o None si no se pudo leer."""
    import xml.etree.ElementTree as ET
    try:
        xml_file.seek(0)
        root = ET.fromstring(xml_file.read())
    except Exception:
        return None

    def _d(v):
        try:
            return decimal.Decimal(v)
        except (TypeError, decimal.InvalidOperation):
            return None

    info = {'subtotal': _d(root.get('SubTotal')), 'total': _d(root.get('Total')),
            'traslados': decimal.Decimal('0'), 'retenidos': decimal.Decimal('0'),
            'uuid': '', 'rfc_receptor': '', 'nombre_receptor': '',
            'serie': root.get('Serie') or '', 'folio': root.get('Folio') or '',
            'moneda': root.get('Moneda') or '', 'fecha': root.get('Fecha') or '',
            'metodo_pago': (root.get('MetodoPago') or '').upper(),
            'conceptos': [],
            # Timbre fiscal (para la representación impresa)
            'sello_cfdi': root.get('Sello') or '', 'rfc_emisor': '', 'nombre_emisor': '',
            'fecha_timbrado': '', 'rfc_prov_certif': '', 'sello_cfd': '',
            'no_cert_sat': '', 'sello_sat': '', 'tfd_version': '1.1'}
    for el in root.iter():
        tag = el.tag.split('}')[-1]
        if tag == 'Emisor':
            info['rfc_emisor'] = (el.get('Rfc') or '').upper()
            info['nombre_emisor'] = el.get('Nombre') or ''
        elif tag == 'Receptor':
            info['rfc_receptor'] = (el.get('Rfc') or '').upper()
            info['nombre_receptor'] = el.get('Nombre') or ''
        elif tag == 'Concepto':
            info['conceptos'].append({
                'descripcion': el.get('Descripcion') or '',
                'cantidad': _d(el.get('Cantidad')) or decimal.Decimal('0'),
                'unidad': el.get('Unidad') or '',
                'valor_unitario': _d(el.get('ValorUnitario')) or decimal.Decimal('0'),
                'importe': _d(el.get('Importe')) or decimal.Decimal('0'),
            })
        elif tag == 'TimbreFiscalDigital':
            info['uuid'] = (el.get('UUID') or '').upper()
            info['fecha_timbrado'] = el.get('FechaTimbrado') or ''
            info['rfc_prov_certif'] = el.get('RfcProvCertif') or ''
            info['sello_cfd'] = el.get('SelloCFD') or ''
            info['no_cert_sat'] = el.get('NoCertificadoSAT') or ''
            info['sello_sat'] = el.get('SelloSAT') or ''
            info['tfd_version'] = el.get('Version') or '1.1'
        elif tag == 'Impuestos':
            t, r = el.get('TotalImpuestosTrasladados'), el.get('TotalImpuestosRetenidos')
            if t is not None:
                info['traslados'] = _d(t) or decimal.Decimal('0')
            if r is not None:
                info['retenidos'] = _d(r) or decimal.Decimal('0')
    return info


def _parsear_zip_cfdis(zip_file, factura):
    """Lee un ZIP y devuelve la lista de CFDIs encontrados (sin guardar nada):
    [{uuid, serie_folio, fecha(date|None), total, tiene_pdf, duplicado, xml_name, pdf_name, error}]."""
    import io
    import zipfile
    zf = zipfile.ZipFile(zip_file)
    pdfs = {}
    for n in zf.namelist():
        if n.lower().endswith('.pdf'):
            stem = n.rsplit('/', 1)[-1].rsplit('.', 1)[0].upper()
            pdfs[stem] = n
    existentes = set(u.upper() for u in factura.cfdis.values_list('uuid', flat=True) if u)
    items = []
    for n in zf.namelist():
        if not n.lower().endswith('.xml'):
            continue
        nombre = n.rsplit('/', 1)[-1]
        try:
            info = _leer_cfdi(io.BytesIO(zf.read(n)))
        except Exception:
            info = None
        if not info or not info.get('uuid'):
            items.append({'error': True, 'xml_name': nombre})
            continue
        uuid_val = info['uuid'].upper()
        base = nombre.rsplit('.', 1)[0].upper()
        pdf_name = pdfs.get(base) or pdfs.get(uuid_val)
        if not pdf_name:
            for stem, pn in pdfs.items():
                if uuid_val in stem:
                    pdf_name = pn
                    break
        c_sub = info['subtotal'] if info['subtotal'] is not None else decimal.Decimal('0')
        total = info['total'] or (c_sub + info['traslados'] - info['retenidos'])
        sf, fecha = _serie_folio_fecha(info)
        items.append({
            'error': False, 'uuid': uuid_val, 'serie_folio': sf, 'fecha': fecha,
            'total': total, 'tiene_pdf': bool(pdf_name), 'duplicado': uuid_val in existentes,
            'xml_name': n, 'pdf_name': pdf_name, 'info': info, 'c_sub': c_sub,
        })
    return zf, items


def _serie_folio_fecha(cfdi):
    """De un CFDI parseado devuelve (serie_folio, fecha date) para prellenar."""
    from datetime import datetime
    serie = (cfdi.get('serie') or '').strip()
    folio = (cfdi.get('folio') or '').strip()
    sf = f"{serie}-{folio}" if (serie and folio) else (folio or serie)
    fecha = None
    f = (cfdi.get('fecha') or '')[:10]
    if f:
        try:
            fecha = datetime.strptime(f, '%Y-%m-%d').date()
        except ValueError:
            fecha = None
    return sf, fecha


def _qr_data_uri(texto):
    """Genera un QR (PNG data URI) para la representación impresa del CFDI."""
    try:
        import io, base64, qrcode
        img = qrcode.make(texto)
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        return 'data:image/png;base64,' + base64.b64encode(buf.getvalue()).decode()
    except Exception as e:
        print(f'[QR] {e}')
        return ''


def _timbre_de_info(info):
    """Construye el dict del timbre fiscal + QR a partir del XML ya parseado."""
    if not info or not info.get('uuid'):
        return None
    cadena = (f"||{info['tfd_version']}|{info['uuid']}|{info['fecha_timbrado']}|"
              f"{info['rfc_prov_certif']}|{info['sello_cfd']}|{info['no_cert_sat']}||")
    total = info.get('total') or decimal.Decimal('0')
    fe = (info.get('sello_cfdi') or '')[-8:]
    qr_url = (
        "https://verificacfdi.facturaelectronica.sat.gob.mx/default.aspx?"
        f"id={info['uuid']}&re={info['rfc_emisor']}&rr={info['rfc_receptor']}"
        f"&tt={total:017.6f}&fe={fe}")

    def _wrap(s, n=80):
        s = s or ''
        return ' '.join(s[i:i + n] for i in range(0, len(s), n))

    return {
        'uuid': info['uuid'], 'fecha_timbrado': info['fecha_timbrado'],
        'sello_cfdi': _wrap(info['sello_cfdi']), 'sello_sat': _wrap(info['sello_sat']),
        'rfc_prov_certif': info['rfc_prov_certif'], 'no_cert_sat': info['no_cert_sat'],
        'cadena': _wrap(cadena), 'qr': _qr_data_uri(qr_url),
    }


def _info_de_cfdi_row(cfdi_row):
    """Parsea el XML de un CfdiCliente. Devuelve dict info o None."""
    if not cfdi_row or not cfdi_row.archivo_xml:
        return None
    try:
        cfdi_row.archivo_xml.open('rb')
        info = _leer_cfdi(cfdi_row.archivo_xml)
        cfdi_row.archivo_xml.close()
        return info
    except Exception:
        return None


def _timbre_para_pdf(factura):
    """Timbre del primer CFDI con XML de la CxC (para el PDF a nivel cuenta)."""
    cfdi_row = factura.cfdis.exclude(archivo_xml='').exclude(archivo_xml=None).first()
    return _timbre_de_info(_info_de_cfdi_row(cfdi_row))


def _render_cfdi_pdf(cfdi_row, empresa):
    """Representación impresa (marca) de UN CFDI, construida desde su propio XML."""
    import io
    from django.template.loader import get_template
    from xhtml2pdf import pisa
    info = _info_de_cfdi_row(cfdi_row)
    if not info:
        return None
    html = get_template('admon_finanzas/cfdi_pdf.html').render({
        'empresa': empresa, 'cfdi': cfdi_row, 'info': info,
        'timbre': _timbre_de_info(info),
    })
    out = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode('UTF-8')), out)
    return None if pdf.err else out.getvalue()


def _img_url_abs(request, field):
    """URL absoluta y pública de un ImageField (para imágenes en correo, que no
    soportan data: URIs y se cortan si van en base64). '' si no hay."""
    if not field:
        return ''
    try:
        return request.build_absolute_uri(field.url)
    except Exception:
        return ''


def _img_data_uri(field):
    """Devuelve un data URI base64 de un ImageField (para correo/PDF), o '' si no hay."""
    if not field:
        return ''
    import base64
    try:
        field.open('rb')
        data = field.read()
        field.close()
        ext = (field.name.rsplit('.', 1)[-1] or 'png').lower()
        mime = 'jpeg' if ext in ('jpg', 'jpeg') else ext
        return f"data:image/{mime};base64," + base64.b64encode(data).decode()
    except Exception:
        return ''


def _render_factura_pdf(factura, empresa):
    """Genera el PDF propio (marca) de la factura. Devuelve bytes o None."""
    import io
    from django.template.loader import get_template
    from xhtml2pdf import pisa
    partidas = factura.pedido.detalles.select_related('producto') if factura.pedido_id else []
    uuid = factura.uuid_cfdi or (factura.cfdis.first().uuid if factura.cfdis.exists() else '')
    cfdi0 = factura.cfdis.first()
    html = get_template('admon_finanzas/factura_pdf.html').render({
        'factura': factura, 'empresa': empresa, 'partidas': partidas, 'uuid': uuid,
        'fecha_cfdi': (cfdi0.fecha if cfdi0 and cfdi0.fecha else factura.fecha_emision),
        'timbre': _timbre_para_pdf(factura),
    })
    out = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode('UTF-8')), out)
    return None if pdf.err else out.getvalue()


def _sync_factura_cfdis(factura):
    """Acumula el desglose de TODOS los CFDI ligados en la CxC (IVA, retenciones,
    total neto). No toca el subtotal (es el objetivo/base del pedido) ni si ya
    hay cobros aplicados."""
    if factura.total_pagado > 0:
        return
    if not factura.cfdis.exists():
        return
    factura.impuestos = factura._suma_cfdis('traslados')
    factura.retenciones = factura._suma_cfdis('retenciones')
    factura.total = factura._suma_cfdis('total')
    factura.save(update_fields=['impuestos', 'retenciones', 'total'])


def _contexto(request):
    if not request.empresa:
        messages.warning(request, "No hay una empresa activa.")
        return None
    if not request.sucursal_activa:
        messages.warning(request, "No hay una sucursal activa. Selecciona una sede arriba.")
        return None
    return request.empresa, request.sucursal_activa


# --------------------------------------------------------------------------
# CUENTAS POR PAGAR
# --------------------------------------------------------------------------
class CuentasPorPagarView(LoginRequiredMixin, View):
    template_name = 'admon_finanzas/cuentas_por_pagar.html'

    def get(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        from admon_empresas import listas
        from django.urls import reverse
        from admon_compras.models import Proveedor

        qs = FacturaProveedor.objects.filter(empresa=empresa).select_related(
            'proveedor', 'moneda', 'orden_compra').prefetch_related('aplicaciones')

        res = listas.construir(
            request, qs,
            placeholder='Folio, UUID, proveedor o notas',
            search_header=('folio', 'uuid_cfdi', 'proveedor__nombre_fiscal',
                           'proveedor__nombre_comercial', 'notas'),
            date_field='fecha_emision',
            exactos={'estado': 'estado', 'proveedor': 'proveedor_id'},
            filtros_ui=[
                {'name': 'estado', 'label': 'Estado', 'tipo': 'select',
                 'opciones': FacturaProveedor.ESTADO_CHOICES},
                {'name': 'proveedor', 'label': 'Proveedor', 'tipo': 'select',
                 'opciones': [(p.id, str(p)) for p in
                              Proveedor.objects.filter(empresa=empresa).order_by('nombre_fiscal')]},
                {'name': 'desde', 'label': 'Desde', 'tipo': 'date'},
                {'name': 'hasta', 'label': 'Hasta', 'tipo': 'date'},
            ],
            sum_fields=('subtotal', 'impuestos', 'total'),
            clear_url=reverse('admon_finanzas:cuentas_por_pagar'),
            export_nombre='cuentas_por_pagar',
            export_order=('-fecha_emision', '-id'),
            export_columnas=[
                ('Folio', 'folio'), ('Proveedor', lambda o: str(o.proveedor)),
                ('UUID', 'uuid_cfdi'), ('Estado', 'get_estado_display'),
                ('Emisión', lambda o: o.fecha_emision.strftime('%d/%m/%Y') if o.fecha_emision else ''),
                ('Total', 'total'), ('Saldo', lambda o: o.saldo),
                ('Moneda', lambda o: o.moneda.codigo if o.moneda else '')],
        )
        if res['export']:
            return res['export']

        # Totales por pagar sobre el resultado filtrado (puede haber varias monedas)
        por_pagar = {}
        for f in res['qs']:
            if f.estado in ('PENDIENTE', 'PARCIAL'):
                cod = f.moneda.codigo
                por_pagar[cod] = por_pagar.get(cod, decimal.Decimal('0')) + f.saldo

        context = {
            'facturas': res['page_obj'], 'page_obj': res['page_obj'],
            'totales': res['totales'], 'lista': res['lista'],
            'resumen_por_pagar': por_pagar,
            'sucursal_activa': sucursal,
            'seccion': 'finanzas',
        }
        return render(request, self.template_name, context)


class RegistrarFacturaView(LoginRequiredMixin, View):
    """Registra una factura del proveedor contra una OC (puede haber varias)."""
    template_name = 'admon_finanzas/factura_form.html'

    def get(self, request, oc_id):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        orden = get_object_or_404(
            OrdenCompra.objects.select_related('proveedor', 'moneda'),
            id=oc_id, empresa=empresa)

        if orden.estado not in ('AUTORIZADO', 'RECIBIDO', 'FINALIZADO'):
            messages.error(request, "Solo se factura sobre órdenes autorizadas.")
            return redirect('admon_compras:orden_detalle', pk=oc_id)

        facturado = orden.facturas.exclude(estado='CANCELADA').aggregate(s=Sum('total'))['s'] or decimal.Decimal('0')

        context = {
            'orden': orden,
            'ya_facturado': facturado,
            'saldo_oc': orden.total - facturado,
            'sucursal_activa': sucursal,
            'seccion': 'finanzas',
        }
        return render(request, self.template_name, context)

    @transaction.atomic
    def post(self, request, oc_id):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        orden = get_object_or_404(OrdenCompra, id=oc_id, empresa=empresa)

        subtotal = decimal.Decimal(request.POST.get('subtotal') or '0')
        impuestos = decimal.Decimal(request.POST.get('impuestos') or '0')
        total = decimal.Decimal(request.POST.get('total') or '0')
        if total <= 0:
            messages.error(request, "El total de la factura debe ser mayor a 0.")
            return redirect('admon_finanzas:registrar_factura', oc_id=oc_id)

        FacturaProveedor.objects.create(
            empresa=empresa,
            orden_compra=orden,
            proveedor=orden.proveedor,
            folio=request.POST.get('folio'),
            uuid_cfdi=request.POST.get('uuid') or None,
            fecha_emision=request.POST.get('fecha_emision'),
            fecha_vencimiento=request.POST.get('fecha_vencimiento') or None,
            moneda=orden.moneda,
            subtotal=subtotal or total,
            impuestos=impuestos,
            total=total,
            notas=request.POST.get('notas'),
            archivo_xml=request.FILES.get('archivo_xml'),
            archivo_pdf=request.FILES.get('archivo_pdf'),
            registrada_por=request.user,
        )
        messages.success(request, f"Factura {request.POST.get('folio')} registrada para {orden.folio}.")
        return redirect('admon_compras:orden_detalle', pk=oc_id)


class FacturaDetalleView(LoginRequiredMixin, View):
    template_name = 'admon_finanzas/factura_detalle.html'

    def get(self, request, pk):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        factura = get_object_or_404(
            FacturaProveedor.objects.select_related('proveedor', 'moneda', 'orden_compra'),
            pk=pk, empresa=empresa)
        context = {
            'factura': factura,
            'aplicaciones': factura.aplicaciones.select_related('pago', 'pago__metodo'),
            'sucursal_activa': sucursal,
            'seccion': 'finanzas',
        }
        return render(request, self.template_name, context)

    def post(self, request, pk):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        factura = get_object_or_404(FacturaProveedor, pk=pk, empresa=empresa)
        if request.POST.get('accion') == 'cancelar' and factura.total_pagado == 0:
            factura.estado = 'CANCELADA'
            factura.save(update_fields=['estado'])
            messages.info(request, f"Factura {factura.folio} cancelada.")
        else:
            messages.error(request, "No se puede cancelar una factura con pagos aplicados.")
        return redirect('admon_finanzas:factura_detalle', pk=pk)


# --------------------------------------------------------------------------
# PAGOS
# --------------------------------------------------------------------------
class RegistrarPagoView(LoginRequiredMixin, View):
    template_name = 'admon_finanzas/pago_form.html'

    def get(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx

        facturas = services.facturas_por_pagar(empresa)
        context = {
            'facturas': facturas,
            'metodos': MetodoPago.objects.filter(empresa=empresa, activo=True),
            'monedas': Moneda.objects.filter(empresa=empresa, activa=True),
            'moneda_base': empresa.moneda_principal,
            'sucursal_activa': sucursal,
            'seccion': 'finanzas',
        }
        return render(request, self.template_name, context)

    def post(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx

        factura_ids = request.POST.getlist('factura_id[]')
        montos = request.POST.getlist('monto_aplicado[]')
        tipos_cambio = request.POST.getlist('tipo_cambio[]')

        aplicaciones = []
        proveedor = None
        for i, fid in enumerate(factura_ids):
            monto = decimal.Decimal(montos[i] or '0')
            if monto <= 0:
                continue
            factura = get_object_or_404(FacturaProveedor, id=fid, empresa=empresa)
            proveedor = factura.proveedor
            aplicaciones.append({
                'factura': factura,
                'monto_aplicado': monto,
                'tipo_cambio': tipos_cambio[i] if i < len(tipos_cambio) else '1',
            })

        moneda = get_object_or_404(Moneda, id=request.POST.get('moneda'), empresa=empresa)
        metodo = MetodoPago.objects.filter(id=request.POST.get('metodo'), empresa=empresa).first()

        try:
            pago = services.registrar_pago(
                empresa=empresa, tipo='EGRESO', proveedor=proveedor,
                fecha=request.POST.get('fecha'), moneda=moneda, metodo=metodo,
                usuario=request.user, aplicaciones=aplicaciones,
                cuenta_banco=request.POST.get('cuenta_banco'),
                referencia=request.POST.get('referencia'),
                notas=request.POST.get('notas'),
            )
        except services.ErrorPago as e:
            messages.error(request, str(e))
            return redirect('admon_finanzas:registrar_pago')

        messages.success(request, f"Pago {pago.folio} registrado por {moneda.simbolo}{pago.monto}.")
        return redirect('admon_finanzas:historial_pagos')


class AdjuntarComplementoView(LoginRequiredMixin, View):
    """Sube el complemento de pago (REP) que el proveedor emite, normalmente
    días después del pago."""
    def post(self, request, pk):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        pago = get_object_or_404(Pago, pk=pk, empresa=empresa)
        pago.uuid_complemento = request.POST.get('uuid_complemento') or None
        pago.fecha_complemento = request.POST.get('fecha_complemento') or None
        if request.FILES.get('complemento_xml'):
            pago.complemento_xml = request.FILES['complemento_xml']
        if request.FILES.get('complemento_pdf'):
            pago.complemento_pdf = request.FILES['complemento_pdf']
        pago.save()
        messages.success(request, f"Complemento de pago adjuntado a {pago.folio}.")
        return redirect('admon_finanzas:historial_pagos')


class AvisarComplementosView(LoginRequiredMixin, View):
    """Avisa por correo al contador qué cobros (PPD ya pagados) siguen sin su
    complemento de pago (REP), para que los genere."""
    def post(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx

        destino = (request.POST.get('email_destino') or empresa.email_contador or '').strip()
        if not destino:
            messages.error(request, "Captura el correo del contador (o guárdalo en Ajustes de empresa).")
            return redirect('admon_finanzas:historial_pagos')

        qs = Pago.objects.filter(empresa=empresa, tipo='INGRESO').select_related(
            'cliente', 'moneda').prefetch_related('aplicaciones__cfdi', 'aplicaciones__factura_cliente')
        ids_seleccionados = request.POST.getlist('pago_id')
        if ids_seleccionados:
            qs = qs.filter(id__in=ids_seleccionados)
        pendientes = [p for p in qs if p.necesita_complemento]
        if not pendientes:
            messages.info(request, "No hay complementos de pago pendientes por avisar en lo seleccionado.")
            return redirect('admon_finanzas:historial_pagos')

        por_cliente = {}
        total = decimal.Decimal('0')
        n_filas = 0
        for p in pendientes:
            # Una fila por CFDI PPD cubierto por este pago, con su monto aplicado
            aplicaciones_ppd = [ap for ap in p.aplicaciones.all()
                                if ap.cfdi_id and ap.cfdi.metodo_pago == 'PPD']
            cliente_nombre = str(p.cliente or '—')
            grupo = por_cliente.setdefault(
                cliente_nombre, {'cliente': cliente_nombre, 'filas': [], 'subtotal': decimal.Decimal('0')})
            for ap in aplicaciones_ppd:
                total += ap.monto_aplicado
                n_filas += 1
                grupo['subtotal'] += ap.monto_aplicado
                grupo['filas'].append({
                    'folio': p.folio, 'fecha': p.fecha.strftime('%d/%m/%Y'), 'uuid': ap.cfdi.uuid,
                    'monto': f"{p.moneda.simbolo}{ap.monto_aplicado:,.2f} {p.moneda.codigo}",
                })
        if not n_filas:
            messages.info(request, "No hay complementos de pago pendientes por avisar en lo seleccionado.")
            return redirect('admon_finanzas:historial_pagos')

        moneda_sim = pendientes[0].moneda.simbolo
        moneda_cod = pendientes[0].moneda.codigo
        grupos = [
            {'cliente': g['cliente'], 'filas': g['filas'],
             'subtotal': f"{moneda_sim}{g['subtotal']:,.2f} {moneda_cod}"}
            for g in por_cliente.values() if g['filas']
        ]
        # Adjunta el comprobante de transferencia/depósito de cada pago incluido,
        # para que el contador tenga con qué cotejar el REP.
        adjuntos = []
        for p in pendientes:
            if p.comprobante:
                try:
                    p.comprobante.open('rb')
                    data = p.comprobante.read()
                    p.comprobante.close()
                    ext = p.comprobante.name.rsplit('.', 1)[-1]
                    adjuntos.append({'filename': f"Comprobante-{p.folio}.{ext}", 'content': data})
                except Exception as e:
                    print(f'[ADJUNTO COMPROBANTE] {e}')

        from admon_empresas.emails import send_html
        ok = send_html(
            subject=f"[iErp] Complementos de pago pendientes — {empresa.nombre_fiscal}",
            template='admon_finanzas/emails/aviso_complementos.html',
            context={
                'empresa': empresa.nombre_fiscal, 'grupos': grupos,
                'total': f"{moneda_sim}{total:,.2f} {moneda_cod}",
                'isotipo_uri': _img_url_abs(request, empresa.isotipo or empresa.logo),
            },
            to=destino,
            attachments=adjuntos,
        )
        if ok:
            messages.success(request, f"Aviso enviado a {destino} ({n_filas} factura(s) pendientes).")
        else:
            messages.error(request, "No se pudo enviar el correo. Intenta de nuevo.")
        return redirect('admon_finanzas:historial_pagos')


class EstadoCuentaClientePDFView(LoginRequiredMixin, View):
    """Estado de cuenta: CxC de un cliente con saldos pendientes.
    Con ?detalle=1 despliega, por cada CxC, los productos del pedido y los
    CFDIs (facturas) ligados."""
    def get(self, request, cliente_id):
        if not request.empresa:
            return redirect('home')
        import io
        from django.template.loader import get_template
        from admon_ventas.models import Cliente
        detalle = request.GET.get('detalle') in ('1', 'true', 'si')
        cliente = get_object_or_404(Cliente, id=cliente_id, empresa=request.empresa)
        facturas = FacturaCliente.objects.filter(
            empresa=request.empresa, cliente=cliente).exclude(estado='CANCELADA').select_related('moneda', 'pedido')
        if detalle:
            facturas = facturas.prefetch_related('pedido__detalles__producto', 'cfdis')
        pendientes = [f for f in facturas if f.saldo > 0]
        total_saldo = sum((f.saldo for f in pendientes), __import__('decimal').Decimal('0'))
        html = get_template('admon_finanzas/estado_cuenta_pdf.html').render({
            'empresa': request.empresa, 'cliente': cliente,
            'facturas': facturas, 'pendientes': pendientes,
            'total_saldo': total_saldo, 'detalle': detalle,
            'hoy': __import__('django.utils.timezone', fromlist=['now']).now(),
        })
        from xhtml2pdf import pisa
        result = io.BytesIO()
        pdf = pisa.pisaDocument(io.BytesIO(html.encode('UTF-8')), result)
        if pdf.err:
            return HttpResponse("Error al generar el PDF", status=400)
        resp = HttpResponse(result.getvalue(), content_type='application/pdf')
        suf = 'Detallado' if detalle else ''
        resp['Content-Disposition'] = f'inline; filename="EdoCuenta{suf}-{cliente.id}.pdf"'
        return resp


class EstadosCuentaView(LoginRequiredMixin, View):
    """Resumen por cliente: facturado, cobrado y saldo (estados de cuenta)."""
    template_name = 'admon_finanzas/estados_cuenta.html'

    def get(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        Z = decimal.Decimal('0')
        q = (request.GET.get('q') or '').strip().lower()

        facturas = FacturaCliente.objects.filter(empresa=empresa).exclude(
            estado='CANCELADA').select_related('cliente').prefetch_related('aplicaciones')
        data = {}
        for f in facturas:
            d = data.setdefault(f.cliente_id, {
                'cliente': f.cliente, 'facturado': Z, 'cobrado': Z, 'saldo': Z,
                'cxc': 0, 'pendientes': 0})
            d['facturado'] += f.total
            d['cobrado'] += f.total_pagado
            d['saldo'] += f.saldo
            d['cxc'] += 1
            if f.saldo > 0:
                d['pendientes'] += 1
        filas = list(data.values())
        if q:
            filas = [x for x in filas if q in (x['cliente'].nombre_fiscal or '').lower()
                     or q in (x['cliente'].nombre_comercial or '').lower()
                     or q in (x['cliente'].rfc or '').lower()]
        filas.sort(key=lambda x: x['saldo'], reverse=True)
        totales = {
            'facturado': sum((x['facturado'] for x in filas), Z),
            'cobrado': sum((x['cobrado'] for x in filas), Z),
            'saldo': sum((x['saldo'] for x in filas), Z),
        }
        return render(request, self.template_name, {
            'filas': filas, 'totales': totales, 'q': request.GET.get('q', ''),
            'sucursal_activa': sucursal, 'seccion': 'finanzas',
        })


def _render_pago_pdf(pago, empresa):
    """Recibo de cobro/pago en PDF (bytes) o None si xhtml2pdf falla."""
    import io
    from django.template.loader import get_template
    from xhtml2pdf import pisa
    html = get_template('admon_finanzas/recibo_pago.html').render({
        'pago': pago,
        'empresa': empresa,
        'aplicaciones': pago.aplicaciones.select_related('factura', 'factura_cliente'),
        'es_ingreso': pago.tipo == Pago.TIPO_INGRESO,
    })
    result = io.BytesIO()
    pdf = pisa.pisaDocument(io.BytesIO(html.encode('UTF-8')), result)
    if pdf.err:
        return None
    return result.getvalue()


class PagoPDFView(LoginRequiredMixin, View):
    """Recibo de cobro (INGRESO) o comprobante de pago (EGRESO)."""
    def get(self, request, pk):
        if not request.empresa:
            return redirect('home')
        pago = get_object_or_404(
            Pago.objects.select_related('proveedor', 'cliente', 'moneda', 'metodo'),
            pk=pk, empresa=request.empresa)
        pdf_bytes = _render_pago_pdf(pago, request.empresa)
        if pdf_bytes is None:
            return HttpResponse("Error al generar el PDF", status=400)
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = f'inline; filename="{pago.folio}.pdf"'
        return resp


class HistorialPagosView(LoginRequiredMixin, View):
    template_name = 'admon_finanzas/historial_pagos.html'

    def get(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        from admon_empresas import listas
        from django.urls import reverse

        qs = Pago.objects.filter(empresa=empresa).select_related(
            'proveedor', 'cliente', 'moneda', 'metodo').prefetch_related(
            'aplicaciones__factura', 'aplicaciones__factura_cliente')

        res = listas.construir(
            request, qs,
            placeholder='Folio, referencia, proveedor o cliente',
            search_header=('folio', 'referencia', 'proveedor__nombre_fiscal',
                           'proveedor__nombre_comercial', 'cliente__nombre_fiscal',
                           'cliente__nombre_comercial'),
            date_field='fecha',
            exactos={'tipo': 'tipo'},
            filtros_ui=[
                {'name': 'tipo', 'label': 'Tipo', 'tipo': 'select',
                 'opciones': Pago.TIPO_CHOICES},
                {'name': 'desde', 'label': 'Desde', 'tipo': 'date'},
                {'name': 'hasta', 'label': 'Hasta', 'tipo': 'date'},
            ],
            sum_fields=('monto',),
            clear_url=reverse('admon_finanzas:historial_pagos'),
            export_nombre='pagos',
            export_order=('-fecha', '-id'),
            export_columnas=[
                ('Folio', 'folio'), ('Tipo', 'get_tipo_display'),
                ('Tercero', lambda o: str(o.cliente or o.proveedor or '')),
                ('Fecha', lambda o: o.fecha.strftime('%d/%m/%Y') if o.fecha else ''),
                ('Referencia', 'referencia'), ('Monto', 'monto'),
                ('Moneda', lambda o: o.moneda.codigo if o.moneda else '')],
        )
        if res['export']:
            return res['export']
        from django.db.models import Sum, Q
        agg = res['qs'].aggregate(
            egreso=Sum('monto', filter=Q(tipo='EGRESO')),
            ingreso=Sum('monto', filter=Q(tipo='INGRESO')))
        Z = decimal.Decimal('0')
        totales_tipo = {
            'egreso': agg['egreso'] or Z,
            'ingreso': agg['ingreso'] or Z,
            'neto': (agg['ingreso'] or Z) - (agg['egreso'] or Z),
        }
        pendientes_complemento = sum(
            1 for p in Pago.objects.filter(empresa=empresa, tipo='INGRESO')
            .prefetch_related('aplicaciones__cfdi') if p.necesita_complemento)

        context = {
            'pagos': res['page_obj'], 'page_obj': res['page_obj'],
            'totales': res['totales'], 'lista': res['lista'],
            'totales_tipo': totales_tipo,
            'pendientes_complemento': pendientes_complemento,
            'empresa_email_contador': empresa.email_contador,
            'sucursal_activa': sucursal,
            'seccion': 'finanzas',
        }
        return render(request, self.template_name, context)


# --------------------------------------------------------------------------
# CUENTAS POR COBRAR (clientes)
# --------------------------------------------------------------------------
class CuentasPorCobrarView(LoginRequiredMixin, View):
    template_name = 'admon_finanzas/cuentas_por_cobrar.html'

    def get(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        from admon_empresas import listas
        from django.urls import reverse
        from admon_ventas.models import Cliente

        qs = FacturaCliente.objects.filter(empresa=empresa).select_related(
            'cliente', 'moneda', 'pedido').prefetch_related('aplicaciones', 'cfdis')

        # Filtro de facturación: "con factura" = tiene UUID/XML único o algún CFDI ligado.
        from django.db.models import Q, Case, When, Value, IntegerField
        facturada_q = Q(uuid_cfdi__gt='') | Q(archivo_xml__gt='') | Q(cfdis__isnull=False)
        f_fact = (request.GET.get('facturacion') or '').strip()
        if f_fact == 'si':
            qs = qs.filter(facturada_q).distinct()
        elif f_fact == 'no':
            qs = qs.exclude(facturada_q)

        # Por defecto ocultamos las canceladas (salvo que se filtren explícitamente)
        if (request.GET.get('estado') or '').strip() != 'CANCELADA':
            qs = qs.exclude(estado='CANCELADA')

        # Orden: primero lo pendiente (cobro/facturación), al final lo ya pagado/cancelado
        qs = qs.annotate(_prioridad=Case(
            When(estado__in=['PENDIENTE', 'PARCIAL'], then=Value(0)),
            When(estado='PAGADA', then=Value(1)),
            default=Value(2), output_field=IntegerField(),
        )).order_by('_prioridad', '-fecha_emision', '-id')

        res = listas.construir(
            request, qs,
            placeholder='Folio, UUID, cliente o notas',
            search_header=('folio', 'uuid_cfdi', 'cliente__nombre_fiscal',
                           'cliente__nombre_comercial', 'notas'),
            date_field='fecha_emision',
            exactos={'estado': 'estado', 'cliente': 'cliente_id'},
            filtros_ui=[
                {'name': 'estado', 'label': 'Estado', 'tipo': 'select',
                 'opciones': FacturaCliente.ESTADO_CHOICES},
                {'name': 'cliente', 'label': 'Cliente', 'tipo': 'select',
                 'opciones': [(c.id, str(c)) for c in
                              Cliente.objects.filter(empresa=empresa).order_by('nombre_fiscal')]},
                {'name': 'facturacion', 'label': 'Facturación', 'tipo': 'select',
                 'opciones': [('si', 'Facturado'), ('no', 'Sin factura')],
                 'todos': 'Todas'},
                {'name': 'desde', 'label': 'Desde', 'tipo': 'date'},
                {'name': 'hasta', 'label': 'Hasta', 'tipo': 'date'},
            ],
            sum_fields=('subtotal', 'impuestos', 'total'),
            clear_url=reverse('admon_finanzas:cuentas_por_cobrar'),
            export_nombre='cuentas_por_cobrar',
            export_order=('-fecha_emision', '-id'),
            export_columnas=[
                ('Folio', 'folio'), ('Cliente', lambda o: str(o.cliente)),
                ('UUID', 'uuid_cfdi'), ('Estado', 'get_estado_display'),
                ('Emisión', lambda o: o.fecha_emision.strftime('%d/%m/%Y') if o.fecha_emision else ''),
                ('Total', 'total'), ('Saldo', lambda o: o.saldo),
                ('Moneda', lambda o: o.moneda.codigo if o.moneda else '')],
        )
        if res['export']:
            return res['export']

        por_cobrar = {}
        Z = decimal.Decimal('0')
        kpi = {'pendiente_facturar': Z, 'por_cobrar_sin_factura': Z,
               'facturado_por_cobrar': Z, 'total_por_cobrar': Z}
        for f in res['qs']:
            abierta = f.estado in ('PENDIENTE', 'PARCIAL')
            if abierta:
                cod = f.moneda.codigo
                por_cobrar[cod] = por_cobrar.get(cod, Z) + f.saldo
                kpi['total_por_cobrar'] += f.saldo
            if f.estado == 'CANCELADA':
                continue
            if not f.esta_facturada:
                # Lo que falta por timbrar (saldo por facturar, considera parciales)
                kpi['pendiente_facturar'] += f.saldo_por_facturar
                if abierta:
                    kpi['por_cobrar_sin_factura'] += f.saldo
            elif abierta:
                kpi['facturado_por_cobrar'] += f.saldo

        context = {
            'facturas': res['page_obj'], 'page_obj': res['page_obj'],
            'totales': res['totales'], 'lista': res['lista'],
            'resumen_por_cobrar': por_cobrar,
            'kpi': kpi,
            'sucursal_activa': sucursal,
            'seccion': 'finanzas',
        }
        return render(request, self.template_name, context)


class ExcelFacturacionView(LoginRequiredMixin, View):
    """Genera un Excel para el contador con el detalle (partida por partida) de
    las CxC seleccionadas que hay que facturar."""

    def get(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, _sucursal = ctx
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.utils import timezone as _tz

        ids = request.GET.getlist('ids')
        cxcs = (FacturaCliente.objects.filter(empresa=empresa, id__in=ids)
                .exclude(estado='CANCELADA')
                .select_related('cliente', 'moneda', 'pedido')
                .prefetch_related('pedido__detalles__producto'))
        if not cxcs:
            messages.warning(request, "Selecciona al menos una cuenta por cobrar para facturar.")
            return redirect('admon_finanzas:cuentas_por_cobrar')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Por facturar"
        cols = ['Cliente', 'RFC', 'CxC', 'Pedido', 'Fecha', 'SKU', 'Descripción',
                'Cantidad', 'P. Unitario', 'Importe', 'IVA %', 'IVA', 'Total']
        ws.append(cols)
        head_fill = PatternFill('solid', fgColor='0F172A')
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF', size=10)
            c.fill = head_fill
            c.alignment = Alignment(horizontal='center')

        D = decimal.Decimal
        tot_imp = tot_iva = tot_tot = D('0')
        for f in cxcs:
            cliente = f.cliente.nombre_fiscal
            rfc = (f.cliente.rfc or '').upper()
            fecha = f.fecha_emision.strftime('%d/%m/%Y') if f.fecha_emision else ''
            dets = list(f.pedido.detalles.all()) if f.pedido_id else []
            # La CxC puede ser una entrega PARCIAL (anticipo/hito). Como no guarda su
            # propio desglose, prorrateamos las partidas del pedido a la proporción que
            # representa esta CxC sobre el total, para que los renglones sumen su importe real.
            ped_sub = sum((d.cantidad * d.precio_unitario for d in dets), D('0'))
            ratio = (f.subtotal / ped_sub) if ped_sub else D('1')
            n = len(dets)
            acc = D('0')   # acumulado para que el último renglón absorba el redondeo
            for i, d in enumerate(dets):
                base = (d.cantidad * d.precio_unitario)
                if i < n - 1:
                    importe = (base * ratio).quantize(D('0.01'))
                else:
                    importe = (f.subtotal - acc).quantize(D('0.01'))
                acc += importe
                # Cantidad consistente con el importe prorrateado (cant × PU = importe)
                cant = (importe / d.precio_unitario) if d.precio_unitario else (d.cantidad * ratio)
                tasa = d.iva_porcentaje or D('0')
                iva = (importe * tasa / 100).quantize(D('0.01'))
                if d.es_retencion:
                    iva = -iva
                total = importe + iva
                ws.append([
                    cliente, rfc, f.folio, f.pedido.folio if f.pedido_id else '', fecha,
                    d.producto.sku, d.producto.nombre,
                    float(round(cant, 4)), float(d.precio_unitario), float(importe),
                    float(tasa), float(iva), float(total)])
                tot_imp += importe; tot_iva += iva; tot_tot += total

        # Fila de totales
        ws.append([])
        fila_tot = ['', '', '', '', '', '', 'TOTALES', '', '', float(tot_imp), '',
                    float(tot_iva), float(tot_tot)]
        ws.append(fila_tot)
        for c in ws[ws.max_row]:
            c.font = Font(bold=True)

        # Anchos y formato de moneda
        anchos = [30, 14, 12, 14, 11, 16, 50, 10, 13, 14, 7, 13, 14]
        for i, w in enumerate(anchos, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for col in (9, 10, 12, 13):
                row[col-1].number_format = '#,##0.00'

        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = (
            f'attachment; filename="por_facturar_{_tz.now():%Y%m%d_%H%M}.xlsx"')
        wb.save(resp)
        return resp


class FacturaClienteDetalleView(LoginRequiredMixin, View):
    template_name = 'admon_finanzas/factura_cliente_detalle.html'

    def get(self, request, pk):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        factura = get_object_or_404(
            FacturaCliente.objects.select_related('cliente', 'moneda', 'pedido'),
            pk=pk, empresa=empresa)
        partidas = factura.pedido.detalles.select_related('producto') if factura.pedido_id else []
        context = {
            'factura': factura,
            'partidas': partidas,
            'cfdis': factura.cfdis.prefetch_related('aplicaciones'),
            'aplicaciones': factura.aplicaciones.select_related('pago', 'pago__metodo'),
            'metodos': MetodoPago.objects.filter(empresa=empresa, activo=True),
            'monedas': Moneda.objects.filter(empresa=empresa, activa=True),
            'sucursal_activa': sucursal,
            'seccion': 'finanzas',
        }
        return render(request, self.template_name, context)

    def post(self, request, pk):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        factura = get_object_or_404(FacturaCliente, pk=pk, empresa=empresa)
        accion = request.POST.get('accion')
        if accion == 'preview_zip':
            from django.http import JsonResponse
            import zipfile
            z = request.FILES.get('archivo_zip')
            if not z:
                return JsonResponse({'error': 'Selecciona un archivo .zip.'}, status=400)
            try:
                _zf, items = _parsear_zip_cfdis(z, factura)
            except zipfile.BadZipFile:
                return JsonResponse({'error': 'El archivo no es un ZIP válido.'}, status=400)
            sim = factura.moneda.simbolo
            out = []
            for it in items:
                if it.get('error'):
                    out.append({'error': True, 'xml_name': it['xml_name']})
                    continue
                out.append({
                    'uuid': it['uuid'], 'serie_folio': it['serie_folio'] or '—',
                    'fecha': it['fecha'].strftime('%d/%m/%Y') if it['fecha'] else '—',
                    'total': f"{sim}{it['total']:,.2f}", 'tiene_pdf': it['tiene_pdf'],
                    'duplicado': it['duplicado'],
                })
            return JsonResponse({'items': out})

        if accion == 'cobrar_cxc':
            try:
                monto = decimal.Decimal(request.POST.get('monto') or '0')
            except decimal.InvalidOperation:
                monto = decimal.Decimal('0')
            if monto <= 0:
                messages.error(request, "Captura un monto mayor a 0.")
                return redirect('admon_finanzas:factura_cliente_detalle', pk=pk)
            moneda = get_object_or_404(Moneda, id=request.POST.get('moneda'), empresa=empresa)
            metodo = MetodoPago.objects.filter(id=request.POST.get('metodo'), empresa=empresa).first()
            try:
                cobro = services.registrar_cobro(
                    empresa=empresa, cliente=factura.cliente,
                    fecha=request.POST.get('fecha') or None, moneda=moneda, metodo=metodo,
                    usuario=request.user, referencia=request.POST.get('referencia'),
                    comprobante=request.FILES.get('comprobante'),
                    aplicaciones=[{'factura': factura, 'monto_aplicado': monto, 'tipo_cambio': '1'}])
                messages.success(request, f"Cobro {cobro.folio} registrado a {factura.folio}.")
            except services.ErrorPago as e:
                messages.error(request, str(e))
            return redirect('admon_finanzas:factura_cliente_detalle', pk=pk)
        if accion == 'cobrar_cfdi':
            cfdi = get_object_or_404(CfdiCliente, id=request.POST.get('cfdi_id'), factura=factura)
            try:
                monto = decimal.Decimal(request.POST.get('monto') or '0')
            except decimal.InvalidOperation:
                monto = decimal.Decimal('0')
            if monto <= 0:
                messages.error(request, "Captura un monto mayor a 0.")
                return redirect('admon_finanzas:factura_cliente_detalle', pk=pk)
            moneda = get_object_or_404(Moneda, id=request.POST.get('moneda'), empresa=empresa)
            metodo = MetodoPago.objects.filter(id=request.POST.get('metodo'), empresa=empresa).first()
            try:
                cobro = services.registrar_cobro(
                    empresa=empresa, cliente=factura.cliente,
                    fecha=request.POST.get('fecha') or None, moneda=moneda, metodo=metodo,
                    usuario=request.user, referencia=request.POST.get('referencia'),
                    comprobante=request.FILES.get('comprobante'),
                    aplicaciones=[{'factura': factura, 'cfdi': cfdi,
                                   'monto_aplicado': monto, 'tipo_cambio': '1'}])
                messages.success(request, f"Cobro {cobro.folio} aplicado a la factura {cfdi.serie_folio or cfdi.uuid[:8]}.")
            except services.ErrorPago as e:
                messages.error(request, str(e))
            return redirect('admon_finanzas:factura_cliente_detalle', pk=pk)
        if accion == 'subir_cfdi':
            xml = request.FILES.get('archivo_xml')
            pdf = request.FILES.get('archivo_pdf')
            uuid = (request.POST.get('uuid_cfdi') or '').strip().upper()
            serie = (request.POST.get('cfdi_serie') or '').strip()
            fecha = request.POST.get('cfdi_fecha') or None

            # Lee el XML (traslados, retenciones, total, UUID) para cuadrar con el SAT
            cfdi = _leer_cfdi(xml) if xml else None
            if cfdi:
                if not uuid and cfdi['uuid']:
                    uuid = cfdi['uuid']
                # Serie/folio y fecha: si no vinieron en el formulario, del XML
                sf_xml, fecha_xml = _serie_folio_fecha(cfdi)
                serie = serie or sf_xml
                if not fecha:
                    fecha = fecha_xml.isoformat() if fecha_xml else None

            try:
                total = decimal.Decimal(request.POST.get('cfdi_total') or '0')
            except decimal.InvalidOperation:
                total = decimal.Decimal('0')
            if total <= 0 and cfdi and cfdi['total']:
                total = cfdi['total']
            if total <= 0:
                total = factura.saldo_por_facturar

            if not uuid and not xml and not pdf:
                messages.error(request, "Sube al menos el XML/PDF o captura el UUID del CFDI.")
                return redirect('admon_finanzas:factura_cliente_detalle', pk=pk)
            uuid_val = uuid or f"SIN-UUID-{factura.cfdis.count() + 1}"
            if factura.cfdis.filter(uuid=uuid_val).exists():
                messages.error(request, f"Ese CFDI ({uuid_val}) ya está ligado a esta CxC.")
                return redirect('admon_finanzas:factura_cliente_detalle', pk=pk)
            # Desglose del CFDI (del XML si está; si no, todo al subtotal)
            c_sub = cfdi['subtotal'] if (cfdi and cfdi['subtotal'] is not None) else None
            c_tras = cfdi['traslados'] if cfdi else decimal.Decimal('0')
            c_ret = cfdi['retenidos'] if cfdi else decimal.Decimal('0')
            if c_sub is None:
                # Sin XML: asumimos que el monto capturado es la base (sin impuestos)
                c_sub = total
            CfdiCliente.objects.create(
                factura=factura, uuid=uuid_val, serie_folio=serie, fecha=fecha,
                subtotal=c_sub, traslados=c_tras, retenciones=c_ret, total=total,
                metodo_pago=(cfdi['metodo_pago'] if cfdi else ''),
                archivo_xml=xml, archivo_pdf=pdf)

            _sync_factura_cfdis(factura)
            factura.refresh_from_db()

            if factura.retenciones > 0:
                messages.success(
                    request,
                    f"CFDI agregado. CxC al SAT: IVA {factura.moneda.simbolo}{factura.impuestos:,.2f}, "
                    f"retenciones {factura.moneda.simbolo}{factura.retenciones:,.2f}, "
                    f"total {factura.moneda.simbolo}{factura.total:,.2f}.")
            elif factura.saldo_por_facturar > 0:
                messages.success(
                    request,
                    f"CFDI agregado. Base facturada {factura.moneda.simbolo}{factura.subtotal_facturado:,.2f} "
                    f"de {factura.moneda.simbolo}{factura.subtotal:,.2f}; faltan "
                    f"{factura.moneda.simbolo}{factura.saldo_por_facturar:,.2f} por facturar.")
            else:
                messages.success(request, "CFDI agregado. La CxC quedó totalmente facturada.")
        elif accion == 'subir_cfdi_zip':
            # Agrega los CFDIs del ZIP que el usuario confirmó en la vista previa.
            import zipfile
            from django.core.files.base import ContentFile
            z = request.FILES.get('archivo_zip')
            if not z:
                messages.error(request, "Selecciona un archivo .zip.")
                return redirect('admon_finanzas:factura_cliente_detalle', pk=pk)
            try:
                zf, items = _parsear_zip_cfdis(z, factura)
            except zipfile.BadZipFile:
                messages.error(request, "El archivo no es un ZIP válido.")
                return redirect('admon_finanzas:factura_cliente_detalle', pk=pk)

            # UUIDs marcados en la vista previa (si no vino ninguno, no agrega nada)
            seleccion = set(u.upper() for u in request.POST.getlist('cfdi_uuids'))
            creados = omitidos = 0
            for it in items:
                if it.get('error') or it['duplicado']:
                    continue
                if it['uuid'] not in seleccion:
                    continue
                c = CfdiCliente(
                    factura=factura, uuid=it['uuid'], serie_folio=it['serie_folio'], fecha=it['fecha'],
                    subtotal=it['c_sub'], traslados=it['info']['traslados'],
                    retenciones=it['info']['retenidos'], total=it['total'],
                    metodo_pago=it['info'].get('metodo_pago', ''))
                c.archivo_xml.save(f"{it['uuid']}.xml", ContentFile(zf.read(it['xml_name'])), save=False)
                if it['pdf_name']:
                    c.archivo_pdf.save(f"{it['uuid']}.pdf", ContentFile(zf.read(it['pdf_name'])), save=False)
                c.save()
                creados += 1

            if not creados:
                messages.info(request, "No se agregó ningún CFDI (nada seleccionado).")
                return redirect('admon_finanzas:factura_cliente_detalle', pk=pk)
            _sync_factura_cfdis(factura)
            factura.refresh_from_db()
            messages.success(request, f"{creados} CFDI agregado(s) desde el ZIP.")
            return redirect('admon_finanzas:factura_cliente_detalle', pk=pk)
        elif accion == 'enviar_cfdi':
            from admon_empresas.emails import send_html
            destino = (request.POST.get('email_destino') or factura.cliente.email or '').strip()
            if not destino:
                messages.error(request, "El cliente no tiene correo. Captura uno para enviar.")
                return redirect('admon_finanzas:factura_cliente_detalle', pk=pk)

            adjuntos = []
            def _add_field(f, fallback):
                if f:
                    try:
                        f.open('rb'); data = f.read(); f.close()
                        adjuntos.append({'filename': f.name.split('/')[-1] or fallback,
                                         'content': data})
                    except Exception as e:
                        print(f'[ADJUNTO CFDI] {e}')

            # CFDI seleccionados (si no marcan ninguno, se envían todos)
            seleccion = set(request.POST.getlist('cfdi_ids'))
            cfdis = list(factura.cfdis.all())
            if seleccion:
                cfdis = [c for c in cfdis if str(c.id) in seleccion]
            n_facturas = 0
            for c in cfdis:
                _add_field(c.archivo_xml, f"{c.uuid}.xml")          # XML fiscal de cada factura
                pdf = _render_cfdi_pdf(c, request.empresa)          # un PDF por cada factura
                if pdf:
                    nombre = f"Factura-{c.serie_folio or c.uuid[:8]}.pdf"
                    adjuntos.append({'filename': nombre, 'content': pdf})
                n_facturas += 1

            # Compat: CxC vieja con XML/PDF únicos (sin filas CfdiCliente)
            if not cfdis and not factura.cfdis.exists():
                _add_field(factura.archivo_xml, 'cfdi.xml')
                pdf_bytes = _render_factura_pdf(factura, request.empresa)
                if pdf_bytes:
                    adjuntos.append({'filename': f"Factura-{factura.folio}.pdf", 'content': pdf_bytes})

            if not adjuntos:
                messages.error(request, "No hay facturas que enviar. Sube el XML del CFDI primero o selecciona al menos una.")
                return redirect('admon_finanzas:factura_cliente_detalle', pk=pk)

            ok = send_html(
                subject=f"Factura {factura.folio} — {request.empresa.nombre_fiscal}",
                template='admon_finanzas/emails/factura_cliente.html',
                context={
                    'empresa': request.empresa.nombre_fiscal,
                    'isotipo_uri': _img_url_abs(request, request.empresa.isotipo or request.empresa.logo),
                    'cliente': str(factura.cliente),
                    'pedido': factura.pedido.folio if factura.pedido_id else '',
                    'folio': factura.folio,
                    'uuid': factura.uuid_cfdi or '',
                    'moneda': factura.moneda.simbolo,
                    'total': f"{factura.total:,.2f}",
                    'mensaje': (request.POST.get('mensaje') or '').strip(),
                },
                to=destino,
                attachments=adjuntos)
            if ok:
                from django.utils import timezone as _tz
                ahora = _tz.now()
                if cfdis:
                    for c in cfdis:
                        c.enviado_en = ahora
                        c.save(update_fields=['enviado_en'])
                else:
                    factura.enviado_en = ahora
                    factura.save(update_fields=['enviado_en'])
                messages.success(request, f"Enviado a {destino}: {len(adjuntos)} archivo(s).")
            else:
                messages.error(request, "No se pudo enviar el correo. Revisa la configuración de correo.")
        elif accion == 'quitar_cfdi_legacy':
            ref = factura.uuid_cfdi or 'CFDI'
            factura.uuid_cfdi = None
            factura.archivo_xml = None
            factura.archivo_pdf = None
            factura.save(update_fields=['uuid_cfdi', 'archivo_xml', 'archivo_pdf'])
            messages.info(request, f"CFDI {ref} quitado. Ya puedes subir el correcto.")
        elif accion == 'eliminar_cfdi':
            cfdi = get_object_or_404(CfdiCliente, id=request.POST.get('cfdi_id'), factura=factura)
            ref = cfdi.uuid
            cfdi.delete()
            _sync_factura_cfdis(factura)
            messages.info(request, f"CFDI {ref} eliminado de la cuenta por cobrar.")
        elif accion == 'cancelar' and factura.total_pagado == 0:
            factura.estado = 'CANCELADA'
            factura.save(update_fields=['estado'])
            messages.info(request, f"CxC {factura.folio} cancelada.")
        else:
            messages.error(request, "No se puede cancelar una factura con cobros aplicados.")
        return redirect('admon_finanzas:factura_cliente_detalle', pk=pk)


class FacturaPDFView(LoginRequiredMixin, View):
    """Descarga el PDF propio (marca) de la factura del cliente."""
    def get(self, request, pk):
        if not request.empresa:
            return redirect('home')
        factura = get_object_or_404(
            FacturaCliente.objects.select_related('cliente', 'moneda', 'pedido'),
            pk=pk, empresa=request.empresa)
        pdf_bytes = _render_factura_pdf(factura, request.empresa)
        if not pdf_bytes:
            return HttpResponse("Error al generar el PDF", status=400)
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = f'inline; filename="Factura-{factura.folio}.pdf"'
        return resp


class CfdiPDFView(LoginRequiredMixin, View):
    """Representación impresa (marca) de UN CFDI específico de la CxC."""
    def get(self, request, pk, cfdi_id):
        if not request.empresa:
            return redirect('home')
        factura = get_object_or_404(FacturaCliente, pk=pk, empresa=request.empresa)
        cfdi = get_object_or_404(CfdiCliente, id=cfdi_id, factura=factura)
        pdf_bytes = _render_cfdi_pdf(cfdi, request.empresa)
        if not pdf_bytes:
            return HttpResponse("No se pudo generar el PDF (¿el CFDI tiene XML?).", status=400)
        resp = HttpResponse(pdf_bytes, content_type='application/pdf')
        resp['Content-Disposition'] = f'inline; filename="Factura-{cfdi.serie_folio or cfdi.uuid[:8]}.pdf"'
        return resp


class RegistrarCobroView(LoginRequiredMixin, View):
    template_name = 'admon_finanzas/cobro_form.html'

    def get(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        context = {
            'facturas': services.facturas_por_cobrar(empresa).prefetch_related(
                'cfdis__aplicaciones', 'aplicaciones'),
            'metodos': MetodoPago.objects.filter(empresa=empresa, activo=True),
            'monedas': Moneda.objects.filter(empresa=empresa, activa=True),
            'moneda_base': empresa.moneda_principal,
            'sucursal_activa': sucursal,
            'seccion': 'finanzas',
        }
        return render(request, self.template_name, context)

    def post(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx

        factura_ids = request.POST.getlist('factura_id[]')
        cfdi_ids = request.POST.getlist('cfdi_id[]')
        montos = request.POST.getlist('monto_aplicado[]')
        tipos_cambio = request.POST.getlist('tipo_cambio[]')

        aplicaciones = []
        cliente = None
        for i, fid in enumerate(factura_ids):
            monto_raw = montos[i] if i < len(montos) else '0'
            try:
                monto = decimal.Decimal(monto_raw or '0')
            except decimal.InvalidOperation:
                monto = decimal.Decimal('0')
            if monto <= 0:
                continue
            factura = get_object_or_404(FacturaCliente, id=fid, empresa=empresa)
            cliente = factura.cliente
            cid = cfdi_ids[i] if i < len(cfdi_ids) else ''
            cfdi = CfdiCliente.objects.filter(id=cid, factura=factura).first() if cid else None
            aplicaciones.append({
                'factura': factura,
                'cfdi': cfdi,
                'monto_aplicado': monto,
                'tipo_cambio': tipos_cambio[i] if i < len(tipos_cambio) else '1',
            })

        moneda = get_object_or_404(Moneda, id=request.POST.get('moneda'), empresa=empresa)
        metodo = MetodoPago.objects.filter(id=request.POST.get('metodo'), empresa=empresa).first()

        try:
            cobro = services.registrar_cobro(
                empresa=empresa, cliente=cliente,
                fecha=request.POST.get('fecha'), moneda=moneda, metodo=metodo,
                usuario=request.user, aplicaciones=aplicaciones,
                cuenta_banco=request.POST.get('cuenta_banco'),
                referencia=request.POST.get('referencia'),
                notas=request.POST.get('notas'),
                comprobante=request.FILES.get('comprobante'),
            )
        except services.ErrorPago as e:
            messages.error(request, str(e))
            return redirect('admon_finanzas:registrar_cobro')

        messages.success(request, f"Cobro {cobro.folio} registrado por {moneda.simbolo}{cobro.monto}.")
        return redirect('admon_finanzas:cuentas_por_cobrar')


# --------------------------------------------------------------------------
# ESTADO DE RESULTADOS (Fase 1: utilidad bruta por periodo)
# --------------------------------------------------------------------------
class EstadoResultadosView(LoginRequiredMixin, View):
    """Estado de resultados simplificado: Ventas netas − Costo de ventas =
    Utilidad bruta, por rango de fechas, con desglose mensual. Base: lo
    entregado (pedidos ENTREGADO/parcial), todo sin IVA."""
    template_name = 'admon_finanzas/estado_resultados.html'

    def get(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        from datetime import date
        from admon_ventas.models import DetallePedido
        D = decimal.Decimal

        hoy = date.today()
        desde = (request.GET.get('desde') or date(hoy.year, 1, 1).isoformat()).strip()
        hasta = (request.GET.get('hasta') or hoy.isoformat()).strip()

        dets = DetallePedido.objects.filter(
            pedido__empresa=empresa,
            pedido__estado__in=['ENTREGADO', 'ENTREGADO_PARCIAL'],
            pedido__fecha_emision__gte=desde,
            pedido__fecha_emision__lte=hasta,
        ).select_related('producto', 'pedido')

        ventas = D('0'); costo = D('0')
        meses = {}  # 'YYYY-MM' -> {ventas, costo}
        for det in dets:
            qty = det.cantidad_entregada or D('0')
            if qty <= 0:
                continue
            v = qty * det.precio_unitario
            c = qty * (det.producto.costo_unitario or D('0'))
            ventas += v; costo += c
            k = det.pedido.fecha_emision.strftime('%Y-%m')
            m = meses.setdefault(k, {'ventas': D('0'), 'costo': D('0')})
            m['ventas'] += v; m['costo'] += c

        util_bruta = ventas - costo
        margen = (util_bruta / ventas * 100) if ventas else D('0')

        # ---- Gastos de operación del periodo, por clasificación y categoría ----
        from .models import Gasto, CategoriaGasto
        gastos = Gasto.objects.filter(
            empresa=empresa, fecha__gte=desde, fecha__lte=hasta
        ).select_related('categoria')
        clasif_labels = dict(CategoriaGasto.CLASIF_CHOICES)
        grupos = {}  # clasif -> {'total': D, 'cats': {nombre: D}}
        total_gastos = D('0')
        for g in gastos:
            clasif = g.categoria.clasificacion
            grp = grupos.setdefault(clasif, {'total': D('0'), 'cats': {}})
            grp['total'] += g.subtotal
            grp['cats'][g.categoria.nombre] = grp['cats'].get(g.categoria.nombre, D('0')) + g.subtotal
            total_gastos += g.subtotal

        gastos_grupos = []
        for clasif, _lbl in CategoriaGasto.CLASIF_CHOICES:
            if clasif in grupos:
                grp = grupos[clasif]
                gastos_grupos.append({
                    'clasificacion': clasif_labels[clasif], 'total': grp['total'],
                    'cats': sorted(grp['cats'].items())})

        util_operacion = util_bruta - total_gastos
        margen_op = (util_operacion / ventas * 100) if ventas else D('0')

        # ---- Partidas no operativas e impuestos (Fase 3) ----
        from .models import PartidaResultado
        partidas = PartidaResultado.objects.filter(
            empresa=empresa, fecha__gte=desde, fecha__lte=hasta)
        otros_ing = D('0'); otros_egr = D('0'); impuestos = D('0')
        for p in partidas:
            if p.naturaleza == 'OTRO_INGRESO':
                otros_ing += p.monto
            elif p.naturaleza == 'OTRO_EGRESO':
                otros_egr += p.monto
            else:
                impuestos += p.monto
        util_antes_imp = util_operacion + otros_ing - otros_egr
        util_neta = util_antes_imp - impuestos
        margen_neto = (util_neta / ventas * 100) if ventas else D('0')

        filas_mes = []
        for k in sorted(meses):
            mv = meses[k]['ventas']; mc = meses[k]['costo']
            filas_mes.append({
                'mes': k, 'ventas': mv, 'costo': mc, 'utilidad': mv - mc,
                'margen': ((mv - mc) / mv * 100) if mv else D('0')})

        # Exportar a Excel (CSV)
        if request.GET.get('export') == 'csv':
            import csv as _csv
            resp = HttpResponse(content_type='text/csv; charset=utf-8-sig')
            resp['Content-Disposition'] = f'attachment; filename="estado_resultados_{desde}_a_{hasta}.csv"'
            resp.write('﻿')
            w = _csv.writer(resp)
            w.writerow(['Concepto', 'Importe'])
            w.writerow(['Ventas netas', ventas.quantize(D('0.01'))])
            w.writerow(['Costo de ventas', (-costo).quantize(D('0.01'))])
            w.writerow(['Utilidad bruta', util_bruta.quantize(D('0.01'))])
            for grp in gastos_grupos:
                w.writerow([f"({grp['clasificacion']})", (-grp['total']).quantize(D('0.01'))])
                for nombre, monto in grp['cats']:
                    w.writerow([f"   {nombre}", (-monto).quantize(D('0.01'))])
            w.writerow(['Total gastos de operación', (-total_gastos).quantize(D('0.01'))])
            w.writerow(['Utilidad de operación', util_operacion.quantize(D('0.01'))])
            if otros_ing:
                w.writerow(['(+) Otros ingresos', otros_ing.quantize(D('0.01'))])
            if otros_egr:
                w.writerow(['(−) Otros gastos', (-otros_egr).quantize(D('0.01'))])
            w.writerow(['Utilidad antes de impuestos', util_antes_imp.quantize(D('0.01'))])
            w.writerow(['(−) Impuestos', (-impuestos).quantize(D('0.01'))])
            w.writerow(['Utilidad neta', util_neta.quantize(D('0.01'))])
            w.writerow([])
            w.writerow(['Mes', 'Ventas netas', 'Costo de ventas', 'Utilidad bruta', 'Margen %'])
            for f in filas_mes:
                w.writerow([f['mes'], f['ventas'].quantize(D('0.01')), f['costo'].quantize(D('0.01')),
                            f['utilidad'].quantize(D('0.01')), f"{f['margen']:.1f}"])
            return resp

        context = {
            'desde': desde, 'hasta': hasta,
            'ventas': ventas, 'costo': costo, 'utilidad': util_bruta, 'margen': margen,
            'gastos_grupos': gastos_grupos, 'total_gastos': total_gastos,
            'util_operacion': util_operacion, 'margen_op': margen_op,
            'otros_ing': otros_ing, 'otros_egr': otros_egr, 'impuestos': impuestos,
            'util_antes_imp': util_antes_imp, 'util_neta': util_neta, 'margen_neto': margen_neto,
            'filas_mes': filas_mes,
            'sucursal_activa': sucursal, 'seccion': 'finanzas',
        }
        return render(request, self.template_name, context)


# --------------------------------------------------------------------------
# GASTOS DE OPERACIÓN (Fase 2)
# --------------------------------------------------------------------------
CATEGORIAS_SEED = [
    ('Comisiones de venta', 'VENTA'), ('Envíos y paquetería', 'VENTA'),
    ('Publicidad y marketing', 'VENTA'),
    ('Renta', 'ADMIN'), ('Servicios (luz, agua, internet)', 'ADMIN'),
    ('Sueldos y honorarios', 'ADMIN'), ('Papelería y oficina', 'ADMIN'),
    ('Software y suscripciones', 'ADMIN'),
    ('Comisiones bancarias', 'FINANCIERO'), ('Intereses', 'FINANCIERO'),
    ('Otros gastos', 'OTRO'),
]


def _seed_categorias(empresa):
    from .models import CategoriaGasto
    if not CategoriaGasto.objects.filter(empresa=empresa).exists():
        CategoriaGasto.objects.bulk_create([
            CategoriaGasto(empresa=empresa, nombre=n, clasificacion=c)
            for n, c in CATEGORIAS_SEED])


class GastosView(LoginRequiredMixin, View):
    template_name = 'admon_finanzas/gastos.html'

    def get(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        from .models import Gasto, CategoriaGasto
        from admon_empresas import listas
        from django.urls import reverse
        _seed_categorias(empresa)

        qs = Gasto.objects.filter(empresa=empresa).select_related('categoria', 'metodo')
        cats = CategoriaGasto.objects.filter(empresa=empresa, activo=True)
        res = listas.construir(
            request, qs,
            placeholder='Descripción, proveedor o referencia',
            search_header=('descripcion', 'proveedor_nombre', 'referencia'),
            date_field='fecha',
            exactos={'categoria': 'categoria_id', 'clasificacion': 'categoria__clasificacion'},
            filtros_ui=[
                {'name': 'categoria', 'label': 'Categoría', 'tipo': 'select',
                 'opciones': [(c.id, c.nombre) for c in cats]},
                {'name': 'clasificacion', 'label': 'Clasificación', 'tipo': 'select',
                 'opciones': CategoriaGasto.CLASIF_CHOICES},
                {'name': 'desde', 'label': 'Desde', 'tipo': 'date'},
                {'name': 'hasta', 'label': 'Hasta', 'tipo': 'date'},
            ],
            sum_fields=('subtotal', 'iva', 'total'),
            clear_url=reverse('admon_finanzas:gastos'),
            export_nombre='gastos', export_order=('-fecha', '-id'),
            export_columnas=[
                ('Fecha', lambda o: o.fecha.strftime('%d/%m/%Y') if o.fecha else ''),
                ('Categoría', lambda o: o.categoria.nombre),
                ('Clasificación', lambda o: o.categoria.get_clasificacion_display()),
                ('Descripción', 'descripcion'), ('Proveedor', 'proveedor_nombre'),
                ('Subtotal', 'subtotal'), ('IVA', 'iva'), ('Total', 'total'),
                ('Referencia', 'referencia')],
        )
        if res['export']:
            return res['export']
        context = {
            'gastos': res['page_obj'], 'page_obj': res['page_obj'],
            'totales': res['totales'], 'lista': res['lista'],
            'categorias': cats,
            'sucursal_activa': sucursal, 'seccion': 'finanzas',
        }
        return render(request, self.template_name, context)

    def post(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        from .models import Gasto, CategoriaGasto
        accion = request.POST.get('accion')

        if accion == 'eliminar':
            g = get_object_or_404(Gasto, id=request.POST.get('gasto_id'), empresa=empresa)
            g.delete()
            messages.success(request, "Gasto eliminado.")
            return redirect('admon_finanzas:gastos')

        if accion == 'nueva_categoria':
            nombre = (request.POST.get('cat_nombre') or '').strip()
            clasif = request.POST.get('cat_clasificacion') or 'ADMIN'
            if nombre:
                CategoriaGasto.objects.get_or_create(
                    empresa=empresa, nombre=nombre, defaults={'clasificacion': clasif})
                messages.success(request, f"Categoría '{nombre}' creada.")
            return redirect('admon_finanzas:gastos')

        # Alta de gasto
        try:
            categoria = get_object_or_404(
                CategoriaGasto, id=request.POST.get('categoria'), empresa=empresa)
            subtotal = decimal.Decimal(request.POST.get('subtotal') or '0')
            iva = decimal.Decimal(request.POST.get('iva') or '0')
            if iva == 0 and request.POST.get('aplica_iva') == 'on':
                iva = (subtotal * decimal.Decimal('0.16')).quantize(decimal.Decimal('0.01'))
            g = Gasto.objects.create(
                empresa=empresa, sucursal=sucursal, categoria=categoria,
                fecha=request.POST.get('fecha'),
                descripcion=(request.POST.get('descripcion') or '').strip(),
                proveedor_nombre=(request.POST.get('proveedor_nombre') or '').strip(),
                subtotal=subtotal, iva=iva, total=subtotal + iva,
                referencia=(request.POST.get('referencia') or '').strip(),
                uuid_cfdi=(request.POST.get('uuid_cfdi') or '').strip() or None,
                comprobante=request.FILES.get('comprobante'),
                creado_por=request.user)
            messages.success(request, f"Gasto registrado: ${g.total}.")
        except Exception as e:
            messages.error(request, f"No se pudo registrar el gasto: {e}")
        return redirect('admon_finanzas:gastos')


# --------------------------------------------------------------------------
# OTROS RESULTADOS E IMPUESTOS (Fase 3)
# --------------------------------------------------------------------------
class OtrosResultadosView(LoginRequiredMixin, View):
    template_name = 'admon_finanzas/otros_resultados.html'

    def get(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        from .models import PartidaResultado
        from admon_empresas import listas
        from django.urls import reverse

        qs = PartidaResultado.objects.filter(empresa=empresa)
        res = listas.construir(
            request, qs,
            placeholder='Concepto o referencia',
            search_header=('concepto', 'referencia'),
            date_field='fecha',
            exactos={'naturaleza': 'naturaleza'},
            filtros_ui=[
                {'name': 'naturaleza', 'label': 'Naturaleza', 'tipo': 'select',
                 'opciones': PartidaResultado.NATURALEZA_CHOICES},
                {'name': 'desde', 'label': 'Desde', 'tipo': 'date'},
                {'name': 'hasta', 'label': 'Hasta', 'tipo': 'date'},
            ],
            sum_fields=('monto',),
            clear_url=reverse('admon_finanzas:otros_resultados'),
            export_nombre='otros_resultados', export_order=('-fecha', '-id'),
            export_columnas=[
                ('Fecha', lambda o: o.fecha.strftime('%d/%m/%Y') if o.fecha else ''),
                ('Naturaleza', 'get_naturaleza_display'), ('Concepto', 'concepto'),
                ('Monto', 'monto'), ('Referencia', 'referencia')],
        )
        if res['export']:
            return res['export']
        context = {
            'partidas': res['page_obj'], 'page_obj': res['page_obj'],
            'totales': res['totales'], 'lista': res['lista'],
            'naturalezas': PartidaResultado.NATURALEZA_CHOICES,
            'sucursal_activa': sucursal, 'seccion': 'finanzas',
        }
        return render(request, self.template_name, context)

    def post(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        from .models import PartidaResultado
        if request.POST.get('accion') == 'eliminar':
            p = get_object_or_404(PartidaResultado, id=request.POST.get('partida_id'), empresa=empresa)
            p.delete()
            messages.success(request, "Partida eliminada.")
            return redirect('admon_finanzas:otros_resultados')
        try:
            PartidaResultado.objects.create(
                empresa=empresa, naturaleza=request.POST.get('naturaleza'),
                fecha=request.POST.get('fecha'),
                concepto=(request.POST.get('concepto') or '').strip(),
                monto=decimal.Decimal(request.POST.get('monto') or '0'),
                referencia=(request.POST.get('referencia') or '').strip(),
                creado_por=request.user)
            messages.success(request, "Partida registrada.")
        except Exception as e:
            messages.error(request, f"No se pudo registrar: {e}")
        return redirect('admon_finanzas:otros_resultados')


# --------------------------------------------------------------------------
# CONCILIACIÓN SAT (CFDI emitidos/recibidos vs sistema)
# --------------------------------------------------------------------------
class ConciliacionSATView(LoginRequiredMixin, View):
    template_name = 'admon_finanzas/conciliacion_sat.html'

    def get(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        from . import conciliacion as conc

        filtro = (request.GET.get('estado') or '').strip()  # '', CONCILIADO, SIN_UUID, FALTANTE
        direccion = (request.GET.get('direccion') or '').strip()
        filas = conc.conciliar(empresa)

        resumen = {'CONCILIADO': 0, 'SIN_UUID': 0, 'FALTANTE': 0,
                   'EMITIDO': 0, 'RECIBIDO': 0}
        for f in filas:
            resumen[f['estado']] += 1
            resumen[f['comp'].direccion] += 1

        if filtro:
            filas = [f for f in filas if f['estado'] == filtro]
        if direccion:
            filas = [f for f in filas if f['comp'].direccion == direccion]

        context = {
            'filas': filas, 'resumen': resumen,
            'filtro': filtro, 'direccion_f': direccion,
            'rfc_empresa': empresa.rfc,
            'sucursal_activa': sucursal, 'seccion': 'finanzas',
        }
        return render(request, self.template_name, context)

    def post(self, request):
        ctx = _contexto(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        from . import conciliacion as conc
        from .models import ComprobanteSAT, Gasto, CategoriaGasto, FacturaCliente, FacturaProveedor
        accion = request.POST.get('accion')

        if accion == 'cargar':
            archivos = request.FILES.getlist('archivos')
            if not archivos:
                messages.error(request, "Selecciona uno o más XML (o un ZIP) del SAT.")
                return redirect('admon_finanzas:conciliacion_sat')
            try:
                r = conc.cargar_comprobantes(archivos, empresa)
            except Exception as e:
                messages.error(request, f"No se pudieron procesar los archivos: {e}")
                return redirect('admon_finanzas:conciliacion_sat')
            messages.success(
                request, f"CFDI leídos: {r['leidos']} ({r['emitidos']} emitidos, "
                f"{r['recibidos']} recibidos, {r['nuevos']} nuevos). "
                f"Ajenos a tu RFC: {r['ajenos']}, no-CFDI: {r['no_cfdi']}.")
            return redirect('admon_finanzas:conciliacion_sat')

        if accion == 'limpiar':
            ComprobanteSAT.objects.filter(empresa=empresa).delete()
            messages.info(request, "Comprobantes SAT borrados (puedes volver a cargar).")
            return redirect('admon_finanzas:conciliacion_sat')

        if accion == 'eliminar_comp':
            ComprobanteSAT.objects.filter(
                id=request.POST.get('comprobante_id'), empresa=empresa).delete()
            messages.info(request, "Comprobante quitado de la conciliación.")
            return redirect('admon_finanzas:conciliacion_sat')

        comp = get_object_or_404(ComprobanteSAT, id=request.POST.get('comprobante_id'), empresa=empresa)

        if accion == 'marcar_uuid':
            # Buscar el documento candidato y ponerle el UUID
            doc = None
            if comp.direccion == 'EMITIDO':
                doc = conc.FacturaClienteCand(empresa, comp)
            else:
                cand = conc.FacturaProvCand(empresa, comp) or conc.GastoCand(empresa, comp)
                doc = cand[0] if cand else None
            if doc:
                doc.uuid_cfdi = comp.uuid
                doc.save(update_fields=['uuid_cfdi'])
                messages.success(request, f"UUID asignado a {doc.folio if hasattr(doc,'folio') else doc}.")
            else:
                messages.error(request, "Ya no se encontró un documento que coincida por monto y fecha.")
            return redirect('admon_finanzas:conciliacion_sat')

        if accion == 'alta_gasto':
            if comp.direccion != 'RECIBIDO':
                messages.error(request, "Solo los CFDI recibidos se dan de alta como gasto.")
                return redirect('admon_finanzas:conciliacion_sat')
            cat, _ = CategoriaGasto.objects.get_or_create(
                empresa=empresa, nombre='Gastos por CFDI', defaults={'clasificacion': 'ADMIN'})
            Gasto.objects.create(
                empresa=empresa, sucursal=sucursal, categoria=cat, fecha=comp.fecha,
                descripcion=(comp.serie_folio and f"CFDI {comp.serie_folio}") or "Gasto CFDI",
                proveedor_nombre=comp.nombre_emisor[:160],
                subtotal=comp.subtotal, iva=comp.iva, total=comp.total,
                uuid_cfdi=comp.uuid, referencia=comp.serie_folio, creado_por=request.user)
            messages.success(request, f"Gasto creado desde el CFDI por ${comp.total} (revisa su categoría en Gastos).")
            return redirect('admon_finanzas:conciliacion_sat')

        return redirect('admon_finanzas:conciliacion_sat')
