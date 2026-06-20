"""
Plantilla y carga masiva de proveedores desde Excel (.xlsx).
La definición de columnas es única para generar la plantilla y para importar,
así el archivo descargado siempre coincide con lo que el importador espera.
"""
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from admon_empresas.models import Moneda
from .models import Proveedor


# (etiqueta de columna, atributo del modelo, obligatorio)
COLUMNAS = [
    ('Nombre comercial', 'nombre_comercial', False),
    ('Razón social (nombre fiscal)', 'nombre_fiscal', True),
    ('RFC', 'rfc', False),
    ('Email', 'email', False),
    ('Teléfono', 'telefono', False),
    ('Celular', 'celular', False),
    ('Dirección', 'direccion', False),
    ('Atención con', 'contacto_nombre', False),
    ('Días crédito', 'dias_credito', False),
    ('Moneda', 'moneda', False),
    ('Activo (Sí/No)', 'activo', False),
]

EJEMPLO = {
    'nombre_comercial': 'DIPROMEDIC',
    'nombre_fiscal': 'DIPROMEDIC, S.A. DE C.V.',
    'rfc': 'DIP080220635',
    'email': 'ventas@dipromedic.com',
    'telefono': '5590004888',
    'celular': '5568177643',
    'direccion': 'Galveston No. 53, Col. Nápoles, Benito Juárez, CDMX, C.P. 03810',
    'contacto_nombre': 'Juan Pérez',
    'dias_credito': 30,
    'moneda': 'MXN',
    'activo': 'Sí',
}


def generar_plantilla(con_ejemplo=True):
    """Devuelve los bytes de un .xlsx de plantilla (encabezados + ejemplo)."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Proveedores'
    FONT = 'Arial'

    hdr = PatternFill('solid', start_color='1F3864')
    hdr_req = PatternFill('solid', start_color='C55A11')
    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = [26, 34, 16, 26, 16, 16, 46, 20, 12, 10, 13]

    for j, (label, field, req) in enumerate(COLUMNAS, start=1):
        c = ws.cell(row=1, column=j, value=label)
        c.font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
        c.fill = hdr_req if req else hdr
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        ws.column_dimensions[c.column_letter].width = widths[j - 1]

    if con_ejemplo:
        for j, (label, field, req) in enumerate(COLUMNAS, start=1):
            c = ws.cell(row=2, column=j, value=EJEMPLO.get(field, ''))
            c.font = Font(name=FONT, size=10, italic=True, color='808080')
            c.alignment = Alignment(vertical='top', wrap_text=(field == 'direccion'))

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30

    # Hoja de instrucciones
    ins = wb.create_sheet('Instrucciones')
    ins.column_dimensions['A'].width = 30
    ins.column_dimensions['B'].width = 72
    t = ins.cell(row=1, column=1, value='Cómo llenar esta plantilla')
    t.font = Font(name=FONT, bold=True, size=13, color='1F3864')
    notas = [
        ('Nombre comercial', 'Nombre corto del proveedor.'),
        ('Razón social', 'OBLIGATORIO. Nombre fiscal. Si no lo tienes, repite el comercial.'),
        ('RFC', 'Opcional. Si se repite un RFC ya existente, se actualiza ese proveedor.'),
        ('Email / Teléfono / Celular', 'Opcionales, datos de contacto.'),
        ('Dirección', 'Opcional, en una sola celda.'),
        ('Atención con', 'Opcional, nombre del contacto.'),
        ('Días crédito', 'Número entero. 0 = de contado.'),
        ('Moneda', 'MXN o USD. Si se deja vacío se usa MXN.'),
        ('Activo (Sí/No)', 'Sí = alta. Vacío también se considera Sí.'),
        ('Borra la fila de ejemplo', 'La fila gris es solo ejemplo: bórrala antes de cargar.'),
    ]
    for i, (a, b) in enumerate(notas, start=3):
        ca = ins.cell(row=i, column=1, value=a); cb = ins.cell(row=i, column=2, value=b)
        ca.font = Font(name=FONT, bold=True, size=10)
        cb.font = Font(name=FONT, size=10)
        ca.alignment = cb.alignment = Alignment(vertical='top', wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _norm_bool(v):
    if v is None or str(v).strip() == '':
        return True
    return str(v).strip().lower() in ('sí', 'si', 's', 'x', '1', 'true', 'verdadero', 'activo')


def importar(archivo, empresa):
    """Lee un .xlsx y da de alta/actualiza proveedores de la empresa.
    Upsert por RFC (si viene) y si no por nombre fiscal. Devuelve resumen."""
    wb = load_workbook(archivo, data_only=True)
    ws = wb['Proveedores'] if 'Proveedores' in wb.sheetnames else wb.active

    # Mapea encabezados reales -> índice de columna (tolerante al orden)
    headers = {}
    for j, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = j
    label_to_field = {label.lower(): field for label, field, _ in COLUMNAS}

    col = {}
    for label_lower, j in headers.items():
        if label_lower in label_to_field:
            col[label_to_field[label_lower]] = j
    if 'nombre_fiscal' not in col:
        raise ValueError("La plantilla no tiene la columna 'Razón social (nombre fiscal)'.")

    monedas = {m.codigo.upper(): m for m in Moneda.objects.filter(empresa=empresa)}

    creados = actualizados = 0
    errores = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def val(field):
            j = col.get(field)
            v = row[j - 1] if j and j - 1 < len(row) else None
            return str(v).strip() if v is not None else ''

        nombre_fiscal = val('nombre_fiscal')
        nombre_comercial = val('nombre_comercial')
        if not nombre_fiscal and not nombre_comercial:
            continue  # fila vacía
        if not nombre_fiscal:
            nombre_fiscal = nombre_comercial
        rfc = val('rfc')

        datos = {
            'nombre_comercial': nombre_comercial or None,
            'rfc': rfc,
            'email': val('email') or None,
            'telefono': val('telefono'),
            'celular': val('celular'),
            'direccion': val('direccion'),
            'contacto_nombre': val('contacto_nombre'),
            'activo': _norm_bool(val('activo')),
        }
        dc = val('dias_credito')
        try:
            datos['dias_credito'] = int(float(dc)) if dc else 0
        except ValueError:
            datos['dias_credito'] = 0
        mon = val('moneda').upper() or 'MXN'
        datos['moneda_predeterminada'] = monedas.get(mon)

        try:
            existente = None
            if rfc:
                existente = Proveedor.objects.filter(empresa=empresa, rfc=rfc).first()
            if not existente:
                existente = Proveedor.objects.filter(empresa=empresa, nombre_fiscal=nombre_fiscal).first()

            if existente:
                for k, v in datos.items():
                    setattr(existente, k, v)
                existente.save()
                actualizados += 1
            else:
                Proveedor.objects.create(empresa=empresa, nombre_fiscal=nombre_fiscal, **datos)
                creados += 1
        except Exception as e:
            errores.append(f"Fila {i} ({nombre_fiscal}): {e}")

    return {'creados': creados, 'actualizados': actualizados, 'errores': errores}
