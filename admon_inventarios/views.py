import decimal

from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.db.models import ProtectedError, Sum, Q

from .models import (
    Almacen, Clase, Grupo, Tipo, Producto, Ubicacion, UnidadMedida,
    Existencia, Lote, NumeroSerie, MovimientoInventario,
    RecepcionMaterial, DetalleRecepcion, SerieRecepcion, Consignante,
)
from .forms import (
    AlmacenForm, ClaseForm, GrupoForm, TipoForm, ProductoForm,
    UbicacionForm, UnidadMedidaForm,
)
from .services import registrar_movimiento, StockInsuficiente
from . import import_productos
from . import import_kits
from django.http import HttpResponse


def _contexto_valido(request):
    """Empresa y sucursal activas; si falta algo regresa None y deja mensaje."""
    if not request.empresa:
        messages.warning(request, "No hay una empresa activa. Crea o selecciona una empresa primero.")
        return None
    if not request.sucursal_activa:
        messages.warning(request, "No hay una sucursal activa. Selecciona una sede en la barra superior.")
        return None
    return request.empresa, request.sucursal_activa


def productos_disponibles_en(empresa, sucursal):
    """Productos visibles en una sucursal: los GLOBAL de la empresa
    más los de alcance SUCURSAL asignados explícitamente a esa sucursal."""
    return Producto.objects.filter(empresa=empresa, activo=True).filter(
        Q(alcance=Producto.ALCANCE_GLOBAL) |
        Q(alcance=Producto.ALCANCE_SUCURSAL,
          config_sucursales__sucursal=sucursal,
          config_sucursales__activo_en_sucursal=True)
    ).distinct()


def _puede_ver_precios(request):
    """Costos y precios solo los ve el dueño / superusuario (no el almacenista)."""
    if request.user.is_superuser:
        return True
    perfil = getattr(request.user, 'perfil', None)
    return bool(perfil and perfil.tipo_usuario == 'OWNER')


def puede_editar_maestro(request):
    """El maestro de productos y catálogos se controla desde la sucursal matriz.
    Superusers y dueños siempre pueden."""
    if request.user.is_superuser:
        return True
    perfil = getattr(request.user, 'perfil', None)
    if perfil and perfil.tipo_usuario == 'OWNER':
        return True
    return bool(request.sucursal_activa and request.sucursal_activa.es_matriz)


class CatalogoProductosView(LoginRequiredMixin, View):
    template_name = 'admon_inventarios/catalogos.html'

    def get(self, request):
        empresa = request.empresa
        sucursal_actual = request.sucursal_activa

        if not empresa:
            messages.warning(request, "No hay una empresa activa. Crea o selecciona una empresa primero.")
            return redirect('home')
        if not sucursal_actual:
            messages.warning(request, "No hay una sucursal activa. Selecciona una sede en la barra superior.")
            return redirect('home')

        active_tab = request.session.pop('active_tab_prod', 'productos')

        context = {
            'almacenes': Almacen.objects.filter(sucursal=sucursal_actual),
            'ubicaciones': Ubicacion.objects.filter(almacen__sucursal=sucursal_actual),
            'productos': Producto.objects.filter(empresa=empresa).select_related(
                'clase', 'grupo', 'tipo', 'unidad_medida', 'ubicacion_defecto', 'grupo__ubicacion_defecto'
            ),
            'clases': Clase.objects.filter(empresa=empresa),
            'grupos': Grupo.objects.filter(empresa=empresa),
            'tipos': Tipo.objects.filter(empresa=empresa),
            'unidades': UnidadMedida.objects.filter(empresa=empresa),

            'form_almacen': AlmacenForm(sucursal=sucursal_actual),
            'form_ubicacion': UbicacionForm(sucursal=sucursal_actual),
            'form_unidad': UnidadMedidaForm(),
            'form_producto': ProductoForm(empresa=empresa, sucursal=sucursal_actual),
            'form_clase': ClaseForm(),
            'form_grupo': GrupoForm(sucursal=sucursal_actual),
            'form_tipo': TipoForm(),

            'active_tab': active_tab,
            'sucursal_activa': sucursal_actual,
            'seccion': 'inventarios',
            'puede_editar_maestro': puede_editar_maestro(request),
            'puede_ver_precios': _puede_ver_precios(request),
            'moneda_simbolo': empresa.moneda_principal.simbolo if empresa.moneda_principal else '$',
        }
        return render(request, self.template_name, context)

    def post(self, request):
        empresa = request.empresa
        sucursal_actual = request.sucursal_activa
        if not empresa or not sucursal_actual:
            return redirect('home')

        action = request.POST.get('action')
        item_id = request.POST.get('item_id')
        tab_destino = 'productos'
        es_maestro = puede_editar_maestro(request)

        # Objetos de maestro (controlados por matriz) vs objetos locales de sucursal
        TIPOS_MAESTRO = {'clase', 'grupo', 'tipo', 'producto', 'unidad'}

        if action == 'delete_item':
            tipo_objeto = request.POST.get('tipo_objeto')

            if tipo_objeto in TIPOS_MAESTRO and not es_maestro:
                messages.error(request, "Solo la sucursal matriz puede eliminar registros del maestro.")
                return redirect('admon_inventarios:catalogos_productos')

            modelos = {
                'clase': Clase, 'grupo': Grupo, 'tipo': Tipo,
                'producto': Producto, 'unidad': UnidadMedida,
                'almacen': Almacen, 'ubicacion': Ubicacion,
            }
            model_class = modelos.get(tipo_objeto)
            if model_class:
                if tipo_objeto == 'ubicacion':
                    obj = get_object_or_404(Ubicacion, id=item_id, almacen__empresa=empresa)
                elif tipo_objeto == 'almacen':
                    obj = get_object_or_404(Almacen, id=item_id, empresa=empresa)
                else:
                    obj = get_object_or_404(model_class, id=item_id, empresa=empresa)

                nombre_display = getattr(obj, 'nombre', getattr(obj, 'codigo', 'el registro'))
                try:
                    obj.delete()
                    messages.success(request, f"'{nombre_display}' se eliminó correctamente.")
                except ProtectedError:
                    messages.error(request, f"No se puede eliminar '{nombre_display}': tiene registros asociados.")

            mapeo_tabs = {
                'ubicacion': 'ubicaciones', 'almacen': 'almacenes',
                'unidad': 'unidades', 'clase': 'clases',
                'grupo': 'grupos', 'tipo': 'tipos', 'producto': 'productos',
            }
            request.session['active_tab_prod'] = mapeo_tabs.get(tipo_objeto, 'productos')
            return redirect('admon_inventarios:catalogos_productos')

        # --- Altas/ediciones de maestro: requieren matriz ---
        botones_maestro = ('btn_clase', 'btn_grupo', 'btn_tipo', 'btn_producto', 'btn_unidad')
        if any(b in request.POST for b in botones_maestro) and not es_maestro:
            messages.error(request, "Solo la sucursal matriz puede modificar el maestro de productos y catálogos.")
            return redirect('admon_inventarios:catalogos_productos')

        if 'btn_almacen' in request.POST:
            instance = Almacen.objects.filter(id=item_id, sucursal=sucursal_actual).first() if item_id else None
            form = AlmacenForm(request.POST, instance=instance, sucursal=sucursal_actual)
            if form.is_valid():
                obj = form.save(commit=False)
                obj.empresa = empresa
                obj.sucursal = sucursal_actual
                obj.save()
                messages.success(request, "Almacén guardado.")
            else:
                messages.error(request, f"Error en almacén: {form.errors.as_text()}")
            tab_destino = 'almacenes'

        elif 'btn_ubicacion' in request.POST:
            instance = Ubicacion.objects.filter(id=item_id, almacen__sucursal=sucursal_actual).first() if item_id else None
            form = UbicacionForm(request.POST, instance=instance, sucursal=sucursal_actual)
            if form.is_valid():
                form.save()
                messages.success(request, "Ubicación guardada.")
            else:
                messages.error(request, f"Error en ubicación: {form.errors.as_text()}")
            tab_destino = 'ubicaciones'

        elif 'btn_clase' in request.POST:
            instance = Clase.objects.filter(id=item_id, empresa=empresa).first() if item_id else None
            form = ClaseForm(request.POST, instance=instance)
            if form.is_valid():
                obj = form.save(commit=False)
                obj.empresa = empresa
                obj.save()
                messages.success(request, "Clase guardada.")
            else:
                messages.error(request, f"Error en clase: {form.errors.as_text()}")
            tab_destino = 'clases'

        elif 'btn_grupo' in request.POST:
            instance = Grupo.objects.filter(id=item_id, empresa=empresa).first() if item_id else None
            form = GrupoForm(request.POST, instance=instance, sucursal=sucursal_actual)
            if form.is_valid():
                obj = form.save(commit=False)
                obj.empresa = empresa
                obj.save()
                messages.success(request, "Grupo guardado.")
            else:
                messages.error(request, f"Error en grupo: {form.errors.as_text()}")
            tab_destino = 'grupos'

        elif 'btn_tipo' in request.POST:
            instance = Tipo.objects.filter(id=item_id, empresa=empresa).first() if item_id else None
            form = TipoForm(request.POST, instance=instance)
            if form.is_valid():
                obj = form.save(commit=False)
                obj.empresa = empresa
                obj.save()
                messages.success(request, "Tipo guardado.")
            else:
                messages.error(request, f"Error en tipo: {form.errors.as_text()}")
            tab_destino = 'tipos'

        elif 'btn_producto' in request.POST:
            instance = Producto.objects.filter(id=item_id, empresa=empresa).first() if item_id else None
            form = ProductoForm(request.POST, instance=instance, empresa=empresa, sucursal=sucursal_actual)
            if form.is_valid():
                obj = form.save(commit=False)
                obj.empresa = empresa
                obj.save()
                messages.success(request, "Producto guardado.")
            else:
                messages.error(request, f"Error en producto: {form.errors.as_text()}")
            tab_destino = 'productos'

        elif 'btn_unidad' in request.POST:
            instance = UnidadMedida.objects.filter(id=item_id, empresa=empresa).first() if item_id else None
            form = UnidadMedidaForm(request.POST, instance=instance)
            if form.is_valid():
                obj = form.save(commit=False)
                obj.empresa = empresa
                obj.save()
                messages.success(request, "Unidad de Medida guardada.")
            else:
                messages.error(request, f"Error en unidad: {form.errors.as_text()}")
            tab_destino = 'unidades'

        request.session['active_tab_prod'] = tab_destino
        return redirect('admon_inventarios:catalogos_productos')


class RecepcionesView(LoginRequiredMixin, View):
    template_name = 'admon_inventarios/recepciones.html'

    def get(self, request):
        ctx = _contexto_valido(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        from admon_empresas import listas
        from django.urls import reverse

        qs = RecepcionMaterial.objects.filter(
            empresa=empresa, sucursal=sucursal
        ).select_related('recibido_por', 'orden_compra').prefetch_related('detalles__producto')

        res = listas.construir(
            request, qs,
            placeholder='Factura, proveedor, OC o producto (SKU / nombre)',
            search_header=('numero_factura', 'remision_proveedor', 'proveedor_nombre',
                           'orden_compra__folio', 'notas'),
            detail_model=DetalleRecepcion,
            detail_search=('producto__sku', 'producto__nombre'),
            date_field='fecha_recepcion',
            clear_url=reverse('admon_inventarios:recepciones'),
            export_nombre='recepciones',
            export_order=('-fecha_recepcion',),
            export_columnas=[
                ('Folio', 'folio'), ('Proveedor', 'proveedor_nombre'),
                ('Factura', 'numero_factura'),
                ('OC', lambda o: o.orden_compra.folio if o.orden_compra else ''),
                ('Partidas', lambda o: o.detalles.count()),
                ('Recibió', lambda o: o.recibido_por.username if o.recibido_por else ''),
                ('Fecha', lambda o: o.fecha_recepcion.strftime('%d/%m/%Y %H:%M') if o.fecha_recepcion else '')],
        )
        if res['export']:
            return res['export']
        context = {
            'historico': res['page_obj'], 'page_obj': res['page_obj'],
            'totales': res['totales'], 'lista': res['lista'],
            'sucursal_activa': sucursal,
            'seccion': 'inventarios',
        }
        return render(request, self.template_name, context)


class NuevaRecepcionDirectaView(LoginRequiredMixin, View):
    template_name = 'admon_inventarios/recepcion_directa.html'

    def get(self, request):
        ctx = _contexto_valido(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx

        ubicaciones = Ubicacion.objects.filter(
            almacen__sucursal=sucursal, activa=True
        ).select_related('almacen').order_by('almacen__nombre', 'codigo')

        if not ubicaciones.exists():
            messages.warning(request, "Debes crear al menos un almacén con ubicaciones para recibir mercancía.")
            request.session['active_tab_prod'] = 'almacenes'
            return redirect('admon_inventarios:catalogos_productos')

        productos = productos_disponibles_en(empresa, sucursal).filter(es_comprable=True)

        context = {
            'productos_json': [
                {
                    'id': p.id, 'sku': p.sku, 'nombre': p.nombre,
                    'es_loteable': 'True' if p.es_loteable else 'False',
                    'es_serializable': 'True' if p.es_serializable else 'False',
                    'costo': str(p.costo_unitario),
                    'ubicacion_defecto': p.ubicacion_defecto_id or (p.grupo.ubicacion_defecto_id if p.grupo else None),
                }
                for p in productos.select_related('grupo')
            ],
            'ubicaciones_json': [
                {'id': u.id, 'codigo': u.codigo, 'almacen': u.almacen.nombre}
                for u in ubicaciones
            ],
            'consignantes': Consignante.objects.filter(empresa=empresa, activo=True),
            'puede_ver_precios': _puede_ver_precios(request),
            'sucursal_activa': sucursal,
            'seccion': 'inventarios',
        }
        return render(request, self.template_name, context)

    @transaction.atomic
    def post(self, request):
        ctx = _contexto_valido(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx

        row_uids = request.POST.getlist('row_uid[]')
        producto_ids = request.POST.getlist('producto_id[]')
        cantidades = request.POST.getlist('cantidad[]')
        costos = request.POST.getlist('costo[]')
        ubicaciones_ids = request.POST.getlist('ubicacion_id[]')

        if not producto_ids:
            messages.error(request, "No se recibieron partidas.")
            return redirect('admon_inventarios:nueva_recepcion')

        # Propiedad de la mercancía recibida (propio o a consignación de un tercero)
        propiedad = request.POST.get('propiedad') or 'PROPIO'
        consignante = None
        if propiedad == 'CONSIGNA':
            consignante = Consignante.objects.filter(
                id=request.POST.get('consignante'), empresa=empresa).first()
            if not consignante:
                messages.error(request, "Selecciona el consignante para mercancía en consignación.")
                return redirect('admon_inventarios:nueva_recepcion')

        recepcion = RecepcionMaterial.objects.create(
            empresa=empresa,
            sucursal=sucursal,
            proveedor_nombre=request.POST.get('proveedor'),
            remision_proveedor=request.POST.get('remision'),
            numero_factura=request.POST.get('factura'),
            recibido_por=request.user,
            notas=request.POST.get('notas'),
        )

        try:
            for i, prod_id in enumerate(producto_ids):
                uid = row_uids[i]
                cantidad = decimal.Decimal(cantidades[i] or '0')
                if cantidad <= 0:
                    continue

                producto = get_object_or_404(Producto, id=prod_id, empresa=empresa)
                ubicacion = get_object_or_404(Ubicacion, id=ubicaciones_ids[i], almacen__sucursal=sucursal)
                # Si el almacenista no captura costo (no tiene permiso), usa el del producto
                costo_raw = costos[i] if i < len(costos) else ''
                costo = decimal.Decimal(costo_raw) if costo_raw else producto.costo_unitario

                # --- Serializables: una existencia por serie ---
                if producto.es_serializable:
                    series_lista = request.POST.getlist(f'series_{uid}[]')
                    if len(series_lista) != int(cantidad):
                        raise ValueError(f"Faltan números de serie para {producto.nombre} "
                                         f"({len(series_lista)} capturados de {int(cantidad)}).")

                    detalle = DetalleRecepcion.objects.create(
                        recepcion=recepcion, producto=producto,
                        cantidad_recibida=cantidad, ubicacion=ubicacion, costo_unitario=costo,
                    )
                    for s_num in series_lista:
                        serie_obj, created = NumeroSerie.objects.get_or_create(
                            producto=producto, serie=s_num.strip(),
                        )
                        if not created and Existencia.objects.filter(
                                producto=producto, serie=serie_obj, cantidad__gt=0).exists():
                            raise ValueError(f"La serie {s_num} ya se encuentra en inventario.")

                        registrar_movimiento(
                            empresa=empresa, sucursal=sucursal, producto=producto,
                            ubicacion=ubicacion, tipo='ENTRADA', origen='RECEPCION',
                            cantidad=1, usuario=request.user, serie=serie_obj,
                            referencia=recepcion.folio, costo_unitario=costo,
                            propiedad=propiedad, consignante=consignante,
                        )
                        serie_obj.estado = 'DISPONIBLE'
                        serie_obj.save()
                        SerieRecepcion.objects.create(detalle=detalle, serie=serie_obj)

                # --- Loteables o normales ---
                else:
                    lote_obj = None
                    if producto.es_loteable:
                        lote_num = request.POST.get(f'lote_num_{uid}')
                        lote_fec = request.POST.get(f'lote_fec_{uid}')
                        if not lote_num:
                            raise ValueError(f"El producto {producto.nombre} requiere número de lote.")
                        lote_obj, _ = Lote.objects.get_or_create(
                            producto=producto, numero_lote=lote_num.strip(),
                            defaults={'fecha_caducidad': lote_fec or None},
                        )

                    DetalleRecepcion.objects.create(
                        recepcion=recepcion, producto=producto,
                        cantidad_recibida=cantidad, ubicacion=ubicacion,
                        lote=lote_obj, costo_unitario=costo,
                    )
                    registrar_movimiento(
                        empresa=empresa, sucursal=sucursal, producto=producto,
                        ubicacion=ubicacion, tipo='ENTRADA', origen='RECEPCION',
                        cantidad=cantidad, usuario=request.user, lote=lote_obj,
                        referencia=recepcion.folio, costo_unitario=costo,
                        propiedad=propiedad, consignante=consignante,
                    )

        except (ValueError, StockInsuficiente) as e:
            transaction.set_rollback(True)
            messages.error(request, f"Error: {e}")
            return redirect('admon_inventarios:nueva_recepcion')

        if not recepcion.detalles.exists():
            transaction.set_rollback(True)
            messages.error(request, "Ninguna partida tenía cantidad válida.")
            return redirect('admon_inventarios:nueva_recepcion')

        messages.success(request, f"Recepción {recepcion.folio} procesada exitosamente.")
        return redirect('admon_inventarios:recepciones')


class ExistenciasView(LoginRequiredMixin, View):
    template_name = 'admon_inventarios/existencias.html'

    def get(self, request):
        ctx = _contexto_valido(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx

        stock = Existencia.objects.filter(
            ubicacion__almacen__sucursal=sucursal,
            producto__empresa=empresa,
            cantidad__gt=0,
        ).values(
            'producto__id', 'producto__nombre', 'producto__sku',
            'producto__unidad_medida__codigo',
        ).annotate(total_stock=Sum('cantidad')).order_by('producto__nombre')

        detalle_existencias = Existencia.objects.filter(
            ubicacion__almacen__sucursal=sucursal,
            producto__empresa=empresa,
            cantidad__gt=0,
        ).select_related('producto', 'ubicacion', 'ubicacion__almacen', 'lote', 'serie')

        context = {
            'stock_resumen': stock,
            'detalle': detalle_existencias,
            'sucursal_activa': sucursal,
            'seccion': 'inventarios',
        }
        return render(request, self.template_name, context)


class DescargarPlantillaProductosView(LoginRequiredMixin, View):
    """Plantilla .xlsx de productos (con ejemplo). Solo superuser."""
    def get(self, request):
        if not request.user.is_superuser:
            messages.error(request, "Solo el administrador puede descargar la plantilla.")
            return redirect('admon_inventarios:catalogos_productos')
        contenido = import_productos.generar_plantilla(con_ejemplo=True)
        resp = HttpResponse(
            contenido,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="plantilla_productos.xlsx"'
        return resp


class ImportarProductosView(LoginRequiredMixin, View):
    """Carga masiva de productos desde Excel. Solo superuser y sucursal matriz."""
    def post(self, request):
        ctx = _contexto_valido(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        if not request.user.is_superuser:
            messages.error(request, "Solo el administrador puede hacer carga masiva.")
            return redirect('admon_inventarios:catalogos_productos')
        if not puede_editar_maestro(request):
            messages.error(request, "El maestro de productos solo se edita desde la sucursal matriz.")
            return redirect('admon_inventarios:catalogos_productos')
        archivo = request.FILES.get('archivo')
        if not archivo or not archivo.name.lower().endswith('.xlsx'):
            messages.error(request, "Selecciona un archivo .xlsx (Excel).")
            return redirect('admon_inventarios:catalogos_productos')
        try:
            res = import_productos.importar(archivo, empresa)
        except Exception as e:
            messages.error(request, f"No se pudo procesar el archivo: {e}")
            return redirect('admon_inventarios:catalogos_productos')
        messages.success(
            request, f"Carga masiva: {res['creados']} nuevos, {res['actualizados']} actualizados.")
        for err in res['errores'][:10]:
            messages.warning(request, err)
        return redirect('admon_inventarios:catalogos_productos')


class DescargarPlantillaKitsView(LoginRequiredMixin, View):
    """Plantilla .xlsx de kits (con ejemplo). Solo superuser."""
    def get(self, request):
        if not request.user.is_superuser:
            messages.error(request, "Solo el administrador puede descargar la plantilla.")
            return redirect('admon_inventarios:kits')
        contenido = import_kits.generar_plantilla(con_ejemplo=True)
        resp = HttpResponse(
            contenido,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="plantilla_kits.xlsx"'
        return resp


class ImportarKitsView(LoginRequiredMixin, View):
    """Carga masiva de kits desde Excel. Solo superuser y sucursal matriz."""
    def post(self, request):
        ctx = _contexto_valido(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        if not request.user.is_superuser:
            messages.error(request, "Solo el administrador puede hacer carga masiva.")
            return redirect('admon_inventarios:kits')
        if not puede_editar_maestro(request):
            messages.error(request, "Los kits se editan desde la sucursal matriz.")
            return redirect('admon_inventarios:kits')
        archivo = request.FILES.get('archivo')
        if not archivo or not archivo.name.lower().endswith('.xlsx'):
            messages.error(request, "Selecciona un archivo .xlsx (Excel).")
            return redirect('admon_inventarios:kits')
        try:
            res = import_kits.importar(archivo, empresa)
        except Exception as e:
            messages.error(request, f"No se pudo procesar el archivo: {e}")
            return redirect('admon_inventarios:kits')
        messages.success(
            request, f"Carga masiva de kits: {res['creados']} nuevos, {res['actualizados']} actualizados.")
        for err in res['errores'][:12]:
            messages.warning(request, err)
        return redirect('admon_inventarios:kits')


class KardexView(LoginRequiredMixin, View):
    """Trazabilidad: historial de movimientos de un producto con saldo corrido.
    Filtros: serie/lote (definen el stream y su saldo), sucursal, tipo y fechas
    (filtros de despliegue). Exporta a Excel con ?formato=xlsx."""
    template_name = 'admon_inventarios/kardex.html'

    def get(self, request):
        ctx = _contexto_valido(request)
        if not ctx:
            return redirect('home')
        empresa, sucursal = ctx
        from .services import TIPOS_ENTRADA
        from admon_empresas.models import Sucursal

        productos = Producto.objects.filter(empresa=empresa, activo=True).order_by('sku')
        producto = None
        pid = request.GET.get('producto')
        if pid:
            producto = Producto.objects.filter(id=pid, empresa=empresa).first()

        f = {
            'serie': request.GET.get('serie') or '',
            'lote': request.GET.get('lote') or '',
            'sucursal': request.GET.get('sucursal') or '',
            'tipo': request.GET.get('tipo') or '',
            'desde': request.GET.get('desde') or '',
            'hasta': request.GET.get('hasta') or '',
        }

        movimientos = []
        saldo_actual = decimal.Decimal('0')
        series = lotes = []
        if producto:
            series = NumeroSerie.objects.filter(producto=producto).order_by('serie')
            lotes = Lote.objects.filter(producto=producto).order_by('numero_lote')

            # Stream que define el saldo: producto + serie/lote/sucursal
            stream = MovimientoInventario.objects.filter(empresa=empresa, producto=producto)
            if f['serie']:
                stream = stream.filter(serie_id=f['serie'])
            if f['lote']:
                stream = stream.filter(lote_id=f['lote'])
            if f['sucursal']:
                stream = stream.filter(sucursal_id=f['sucursal'])
            stream = stream.select_related('sucursal', 'ubicacion', 'ubicacion__almacen',
                                           'lote', 'serie', 'consignante', 'usuario').order_by('fecha', 'id')

            saldo = decimal.Decimal('0')
            filas = []
            for m in stream:
                signo = 1 if m.tipo in TIPOS_ENTRADA else -1
                m.entrada = m.cantidad if signo == 1 else None
                m.salida = m.cantidad if signo == -1 else None
                saldo += signo * m.cantidad
                m.saldo = saldo
                filas.append(m)
            saldo_actual = saldo

            # Filtros de despliegue (no afectan el saldo)
            def visible(m):
                if f['tipo'] and m.tipo != f['tipo']:
                    return False
                fch = m.fecha.date().isoformat()
                if f['desde'] and fch < f['desde']:
                    return False
                if f['hasta'] and fch > f['hasta']:
                    return False
                return True
            filas = [m for m in filas if visible(m)]
            filas.reverse()
            movimientos = filas

            if request.GET.get('formato') == 'xlsx':
                return self._export_xlsx(producto, movimientos)

        context = {
            'productos': productos, 'producto': producto,
            'movimientos': movimientos, 'saldo_actual': saldo_actual,
            'series': series, 'lotes': lotes, 'f': f,
            'sucursales': Sucursal.objects.filter(empresa=empresa),
            'tipos': MovimientoInventario.TIPO_CHOICES,
            'sucursal_activa': sucursal, 'seccion': 'inventarios',
        }
        return render(request, self.template_name, context)

    def _export_xlsx(self, producto, movimientos):
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook(); ws = wb.active; ws.title = 'Kardex'
        cols = ['Fecha', 'Movimiento', 'Referencia', 'Sucursal', 'Ubicación', 'Lote', 'Serie',
                'Propiedad', 'Entrada', 'Salida', 'Saldo', 'Usuario']
        for j, c in enumerate(cols, 1):
            cell = ws.cell(row=1, column=j, value=c)
            cell.font = Font(bold=True, color='FFFFFF'); cell.fill = PatternFill('solid', start_color='1F3864')
        for m in reversed(movimientos):  # cronológico en el Excel
            ws.append([
                m.fecha.strftime('%d/%m/%Y %H:%M'), m.get_tipo_display(), m.referencia or '',
                m.sucursal.nombre, f"{m.ubicacion.almacen.codigo}/{m.ubicacion.codigo}",
                m.lote.numero_lote if m.lote else '', m.serie.serie if m.serie else '',
                'Consigna' if m.propiedad == 'CONSIGNA' else 'Propio',
                float(m.entrada) if m.entrada else '', float(m.salida) if m.salida else '',
                float(m.saldo), m.usuario.username,
            ])
        buf = io.BytesIO(); wb.save(buf)
        resp = HttpResponse(buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="Kardex-{producto.sku}.xlsx"'
        return resp
