"""
Plantilla y carga masiva de productos desde Excel (.xlsx).
Catálogos (Clase/Grupo/Tipo/Unidad) se referencian por su CÓDIGO; el impuesto
por su nombre. Si no existen en la empresa, se deja vacío y se reporta aviso.
"""
import io
import decimal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from admon_empresas.models import Impuesto
from .models import Producto, Clase, Grupo, Tipo, UnidadMedida


# (etiqueta, clave interna, obligatorio)
COLUMNAS = [
    ('SKU', 'sku', True),
    ('Nombre', 'nombre', True),
    ('Descripción', 'descripcion', False),
    ('Código de barras', 'codigo_barras', False),
    ('Clase (código)', 'clase', False),
    ('Grupo (código)', 'grupo', False),
    ('Tipo (código)', 'tipo', False),
    ('Unidad (código)', 'unidad', False),
    ('Alcance (GLOBAL/SUCURSAL)', 'alcance', False),
    ('Loteable (Sí/No)', 'es_loteable', False),
    ('Serializable (Sí/No)', 'es_serializable', False),
    ('Comprable (Sí/No)', 'es_comprable', False),
    ('Vendible (Sí/No)', 'es_vendible', False),
    ('Materia prima (Sí/No)', 'es_materia_prima', False),
    ('Producible (Sí/No)', 'es_producible', False),
    ('Retornable (Sí/No)', 'es_retornable', False),
    ('Impuesto (nombre)', 'impuesto', False),
    ('Costo unitario', 'costo_unitario', False),
    ('Precio venta', 'precio_venta', False),
    ('Precio renta', 'precio_renta', False),
    ('Activo (Sí/No)', 'activo', False),
]

EJEMPLO = {
    'sku': 'TORN-4.5', 'nombre': 'Tornillo cortical 4.5mm',
    'descripcion': 'Acero quirúrgico', 'codigo_barras': '7501234567890',
    'clase': 'IMP', 'grupo': 'OST', 'tipo': 'CONS', 'unidad': 'PZA',
    'alcance': 'GLOBAL',
    'es_loteable': 'Sí', 'es_serializable': 'No', 'es_comprable': 'Sí',
    'es_vendible': 'Sí', 'es_materia_prima': 'No', 'es_producible': 'No',
    'es_retornable': 'No', 'impuesto': 'IVA 16%',
    'costo_unitario': 350, 'precio_venta': 900, 'precio_renta': 0, 'activo': 'Sí',
}

# Default de cada flag booleano cuando la celda viene vacía
_BOOL_DEFAULTS = {
    'es_loteable': False, 'es_serializable': False, 'es_comprable': True,
    'es_vendible': True, 'es_materia_prima': False, 'es_producible': False,
    'es_retornable': False, 'activo': True,
}


def _bool(v, default):
    s = str(v).strip().lower() if v is not None else ''
    if s == '':
        return default
    return s in ('sí', 'si', 's', 'x', '1', 'true', 'verdadero')


def _dec(v):
    s = str(v).strip() if v is not None else ''
    if not s:
        return decimal.Decimal('0')
    try:
        return decimal.Decimal(s)
    except Exception:
        return decimal.Decimal('0')


def generar_plantilla(con_ejemplo=True):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Productos'
    FONT = 'Arial'
    hdr = PatternFill('solid', start_color='1F3864')
    hdr_req = PatternFill('solid', start_color='C55A11')
    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for j, (label, key, req) in enumerate(COLUMNAS, start=1):
        c = ws.cell(row=1, column=j, value=label)
        c.font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
        c.fill = hdr_req if req else hdr
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        ws.column_dimensions[c.column_letter].width = 16 if key not in ('nombre', 'descripcion') else 30

    if con_ejemplo:
        for j, (label, key, req) in enumerate(COLUMNAS, start=1):
            c = ws.cell(row=2, column=j, value=EJEMPLO.get(key, ''))
            c.font = Font(name=FONT, size=10, italic=True, color='808080')

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30

    ins = wb.create_sheet('Instrucciones')
    ins.column_dimensions['A'].width = 30
    ins.column_dimensions['B'].width = 74
    ins.cell(row=1, column=1, value='Cómo llenar esta plantilla').font = Font(name=FONT, bold=True, size=13, color='1F3864')
    notas = [
        ('SKU', 'OBLIGATORIO. Código interno único. Si se repite, se actualiza ese producto.'),
        ('Nombre', 'OBLIGATORIO.'),
        ('Clase/Grupo/Tipo/Unidad', 'Usa el CÓDIGO del catálogo (ej. PZA, KG). Si no existe en la empresa, se deja vacío.'),
        ('Alcance', 'GLOBAL (todas las sucursales) o SUCURSAL. Vacío = GLOBAL.'),
        ('Columnas Sí/No', 'Escribe Sí o No. Comprable y Vendible son Sí por defecto; el resto No.'),
        ('Impuesto', 'Nombre del impuesto tal como está en Configuración (ej. "IVA 16%"). Vacío = sin impuesto / default.'),
        ('Precio renta', 'Solo aplica a productos retornables (instrumental que se presta).'),
        ('Fila de ejemplo', 'Bórrala antes de cargar.'),
    ]
    for i, (a, b) in enumerate(notas, start=3):
        ca = ins.cell(row=i, column=1, value=a); cb = ins.cell(row=i, column=2, value=b)
        ca.font = Font(name=FONT, bold=True, size=10); cb.font = Font(name=FONT, size=10)
        ca.alignment = cb.alignment = Alignment(vertical='top', wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def importar(archivo, empresa):
    wb = load_workbook(archivo, data_only=True)
    ws = wb['Productos'] if 'Productos' in wb.sheetnames else wb.active

    headers = {}
    for j, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = j
    label_to_key = {label.lower(): key for label, key, _ in COLUMNAS}
    col = {label_to_key[k]: j for k, j in headers.items() if k in label_to_key}
    if 'sku' not in col or 'nombre' not in col:
        raise ValueError("La plantilla debe tener al menos las columnas 'SKU' y 'Nombre'.")

    clases = {c.codigo.upper(): c for c in Clase.objects.filter(empresa=empresa)}
    grupos = {g.codigo.upper(): g for g in Grupo.objects.filter(empresa=empresa)}
    tipos = {t.codigo.upper(): t for t in Tipo.objects.filter(empresa=empresa)}
    unidades = {u.codigo.upper(): u for u in UnidadMedida.objects.filter(empresa=empresa)}
    impuestos = {i.nombre.strip().lower(): i for i in Impuesto.objects.filter(empresa=empresa)}

    creados = actualizados = 0
    errores = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def val(key):
            j = col.get(key)
            v = row[j - 1] if j and j - 1 < len(row) else None
            return str(v).strip() if v is not None else ''

        sku = val('sku')
        nombre = val('nombre')
        if not sku and not nombre:
            continue
        if not sku or not nombre:
            errores.append(f"Fila {i}: falta SKU o Nombre.")
            continue

        alcance = (val('alcance').upper() or 'GLOBAL')
        if alcance not in ('GLOBAL', 'SUCURSAL'):
            alcance = 'GLOBAL'

        datos = {
            'nombre': nombre,
            'descripcion': val('descripcion') or None,
            'codigo_barras': val('codigo_barras') or None,
            'clase': clases.get(val('clase').upper()) if val('clase') else None,
            'grupo': grupos.get(val('grupo').upper()) if val('grupo') else None,
            'tipo': tipos.get(val('tipo').upper()) if val('tipo') else None,
            'unidad_medida': unidades.get(val('unidad').upper()) if val('unidad') else None,
            'alcance': alcance,
            'es_loteable': _bool(val('es_loteable'), _BOOL_DEFAULTS['es_loteable']),
            'es_serializable': _bool(val('es_serializable'), _BOOL_DEFAULTS['es_serializable']),
            'es_comprable': _bool(val('es_comprable'), _BOOL_DEFAULTS['es_comprable']),
            'es_vendible': _bool(val('es_vendible'), _BOOL_DEFAULTS['es_vendible']),
            'es_materia_prima': _bool(val('es_materia_prima'), _BOOL_DEFAULTS['es_materia_prima']),
            'es_producible': _bool(val('es_producible'), _BOOL_DEFAULTS['es_producible']),
            'es_retornable': _bool(val('es_retornable'), _BOOL_DEFAULTS['es_retornable']),
            'impuesto': impuestos.get(val('impuesto').strip().lower()) if val('impuesto') else None,
            'costo_unitario': _dec(val('costo_unitario')),
            'precio_venta': _dec(val('precio_venta')),
            'precio_renta': _dec(val('precio_renta')),
            'activo': _bool(val('activo'), _BOOL_DEFAULTS['activo']),
        }
        try:
            existente = Producto.objects.filter(empresa=empresa, sku=sku).first()
            if existente:
                for k, v in datos.items():
                    setattr(existente, k, v)
                existente.save()
                actualizados += 1
            else:
                Producto.objects.create(empresa=empresa, sku=sku, **datos)
                creados += 1
        except Exception as e:
            errores.append(f"Fila {i} ({sku}): {e}")

    return {'creados': creados, 'actualizados': actualizados, 'errores': errores}
