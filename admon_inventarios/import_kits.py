"""
Plantilla y carga masiva de kits (plantillas de caja) desde Excel (.xlsx).
Formato: UNA fila por componente; el código de kit se repite. Los productos
se referencian por SKU. Reimportar un kit reemplaza sus componentes.
"""
import io
import decimal
from collections import OrderedDict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import Kit, DetalleKit, Producto


# (etiqueta, clave, obligatorio)
COLUMNAS = [
    ('Kit (código)', 'kit_codigo', True),
    ('Kit (nombre)', 'kit_nombre', False),
    ('SKU producto', 'sku', True),
    ('Cantidad', 'cantidad', True),
    ('Retornable (Sí/No)', 'retornable', False),
    ('Notas', 'notas', False),
]

EJEMPLO = [
    {'kit_codigo': 'KIT-TRAUMA', 'kit_nombre': 'Set de Trauma básico', 'sku': 'SETT',
     'cantidad': 1, 'retornable': 'No', 'notas': ''},
    {'kit_codigo': 'KIT-TRAUMA', 'kit_nombre': 'Set de Trauma básico', 'sku': 'PERFO',
     'cantidad': 1, 'retornable': 'Sí', 'notas': 'Renta de perforador'},
]

_WIDTHS = [16, 30, 18, 12, 16, 28]


def generar_plantilla(con_ejemplo=True):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Kits'
    FONT = 'Arial'
    hdr = PatternFill('solid', start_color='1F3864')
    hdr_req = PatternFill('solid', start_color='C55A11')
    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for j, (label, key, req) in enumerate(COLUMNAS, start=1):
        c = ws.cell(row=1, column=j, value=label)
        c.font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
        c.fill = hdr_req if req else hdr
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        ws.column_dimensions[c.column_letter].width = _WIDTHS[j - 1]

    if con_ejemplo:
        for i, ej in enumerate(EJEMPLO, start=2):
            for j, (label, key, req) in enumerate(COLUMNAS, start=1):
                c = ws.cell(row=i, column=j, value=ej.get(key, ''))
                c.font = Font(name=FONT, size=10, italic=True, color='808080')

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30

    ins = wb.create_sheet('Instrucciones')
    ins.column_dimensions['A'].width = 30
    ins.column_dimensions['B'].width = 74
    ins.cell(row=1, column=1, value='Cómo llenar esta plantilla').font = Font(name=FONT, bold=True, size=13, color='1F3864')
    notas = [
        ('Una fila por componente', 'Repite el código del kit en cada producto que lo compone.'),
        ('Kit (código)', 'OBLIGATORIO. Identificador del kit (ej. KIT-001). Si ya existe, se reemplazan sus componentes.'),
        ('Kit (nombre)', 'Nombre del kit. Toma el de la primera fila de cada kit.'),
        ('SKU producto', 'OBLIGATORIO. Debe existir en el catálogo de productos. Si no existe, se omite esa fila.'),
        ('Cantidad', 'Cantidad estándar del producto en el kit.'),
        ('Retornable (Sí/No)', 'Sí = instrumental/préstamo que regresa (se cobra como renta). No = consumible.'),
        ('Fila(s) de ejemplo', 'Bórralas antes de cargar.'),
    ]
    for i, (a, b) in enumerate(notas, start=3):
        ca = ins.cell(row=i, column=1, value=a); cb = ins.cell(row=i, column=2, value=b)
        ca.font = Font(name=FONT, bold=True, size=10); cb.font = Font(name=FONT, size=10)
        ca.alignment = cb.alignment = Alignment(vertical='top', wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _bool(v):
    return str(v).strip().lower() in ('sí', 'si', 's', 'x', '1', 'true', 'verdadero')


def importar(archivo, empresa):
    wb = load_workbook(archivo, data_only=True)
    ws = wb['Kits'] if 'Kits' in wb.sheetnames else wb.active

    headers = {}
    for j, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = j
    label_to_key = {label.lower(): key for label, key, _ in COLUMNAS}
    col = {label_to_key[k]: j for k, j in headers.items() if k in label_to_key}
    if 'kit_codigo' not in col or 'sku' not in col:
        raise ValueError("La plantilla debe tener 'Kit (código)' y 'SKU producto'.")

    productos = {p.sku: p for p in Producto.objects.filter(empresa=empresa)}

    # Agrupa filas por kit, preservando orden
    grupos = OrderedDict()
    for row in ws.iter_rows(min_row=2, values_only=True):
        def val(key):
            j = col.get(key)
            v = row[j - 1] if j and j - 1 < len(row) else None
            return str(v).strip() if v is not None else ''
        kc = val('kit_codigo')
        sku = val('sku')
        if not kc and not sku:
            continue
        if not kc:
            continue
        g = grupos.setdefault(kc, {'nombre': '', 'comps': []})
        if val('kit_nombre') and not g['nombre']:
            g['nombre'] = val('kit_nombre')
        if sku:
            g['comps'].append({'sku': sku, 'cantidad': val('cantidad'),
                               'retornable': val('retornable'), 'notas': val('notas')})

    creados = actualizados = 0
    errores = []
    for kc, g in grupos.items():
        try:
            kit = Kit.objects.filter(empresa=empresa, codigo=kc).first()
            if kit:
                kit.nombre = g['nombre'] or kit.nombre
                kit.activo = True
                kit.save()
                kit.componentes.all().delete()
                actualizados += 1
            else:
                kit = Kit.objects.create(empresa=empresa, codigo=kc, nombre=g['nombre'] or kc)
                creados += 1
            for comp in g['comps']:
                prod = productos.get(comp['sku'])
                if not prod:
                    errores.append(f"Kit {kc}: SKU '{comp['sku']}' no existe en el catálogo (omitido).")
                    continue
                try:
                    cant = decimal.Decimal(comp['cantidad'] or '1')
                except Exception:
                    cant = decimal.Decimal('1')
                DetalleKit.objects.create(
                    kit=kit, producto=prod, cantidad_requerida=cant,
                    es_retornable=_bool(comp['retornable']), notas=comp['notas'] or None)
        except Exception as e:
            errores.append(f"Kit {kc}: {e}")

    return {'creados': creados, 'actualizados': actualizados, 'errores': errores}
