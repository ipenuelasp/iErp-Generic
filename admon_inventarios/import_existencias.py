"""
Carga inicial de EXISTENCIAS (inventario de apertura) desde Excel (.xlsx).

Sube el stock previo de una empresa que ya existía antes del sistema. Cada fila
es una entrada de inventario (movimiento AJUSTE_POS, referencia "CARGA INICIAL")
con su almacén/ubicación, lote/caducidad y serie según aplique. El costo puede
ir estimado y corregirse después: no afecta lo que se le cobra al cliente.
"""
import io
import decimal
import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .models import Producto, Almacen, Ubicacion, Lote, NumeroSerie
from . import services


# (etiqueta, clave interna, obligatorio)
COLUMNAS = [
    ('SKU', 'sku', True),
    ('Almacén (código)', 'almacen', True),
    ('Ubicación (código)', 'ubicacion', False),
    ('Cantidad', 'cantidad', True),
    ('Costo unitario', 'costo', False),
    ('Lote', 'lote', False),
    ('Caducidad (dd/mm/aaaa)', 'caducidad', False),
    ('Serie', 'serie', False),
]

EJEMPLO = {
    'sku': 'TORN-4.5', 'almacen': 'PRIN', 'ubicacion': 'GEN', 'cantidad': 20,
    'costo': 350, 'lote': 'L-2026-01', 'caducidad': '31/12/2027', 'serie': '',
}


def _dec(v):
    s = str(v).strip() if v is not None else ''
    if not s:
        return None
    try:
        return decimal.Decimal(s.replace(',', ''))
    except Exception:
        return None


def _fecha(v):
    if v is None or str(v).strip() == '':
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.date() if isinstance(v, datetime.datetime) else v
    s = str(v).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def generar_plantilla(con_ejemplo=True):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Existencias'
    FONT = 'Arial'
    hdr = PatternFill('solid', start_color='1F3864')
    hdr_req = PatternFill('solid', start_color='C55A11')
    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for j, (label, key, req) in enumerate(COLUMNAS, start=1):
        c = ws.cell(row=1, column=j, value=label)
        c.font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
        c.fill = hdr_req if req else hdr
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        ws.column_dimensions[c.column_letter].width = 20

    if con_ejemplo:
        for j, (label, key, req) in enumerate(COLUMNAS, start=1):
            ws.cell(row=2, column=j, value=EJEMPLO.get(key, '')).font = Font(
                name=FONT, size=10, italic=True, color='808080')

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30

    ins = wb.create_sheet('Instrucciones')
    ins.column_dimensions['A'].width = 26
    ins.column_dimensions['B'].width = 78
    ins.cell(row=1, column=1, value='Cómo llenar esta plantilla').font = Font(
        name=FONT, bold=True, size=13, color='1F3864')
    notas = [
        ('SKU', 'OBLIGATORIO. Debe existir ya en el catálogo de productos.'),
        ('Almacén (código)', 'OBLIGATORIO. Código del almacén. Si no existe, se crea automáticamente en esta sucursal.'),
        ('Ubicación (código)', 'Opcional. Si se deja vacío se usa "GEN". Se crea si no existe.'),
        ('Cantidad', 'OBLIGATORIO. Piezas a cargar. En productos con SERIE debe ser 1 por fila.'),
        ('Costo unitario', 'Opcional. Puede ir ESTIMADO; se corrige después. No afecta el precio de venta.'),
        ('Lote', 'Solo productos loteables. Se crea si no existe.'),
        ('Caducidad', 'Opcional, del lote. Formato dd/mm/aaaa.'),
        ('Serie', 'Solo productos serializados. Una fila por serie, con Cantidad = 1.'),
        ('Fila de ejemplo', 'Bórrala antes de cargar.'),
    ]
    for i, (a, b) in enumerate(notas, start=3):
        ca = ins.cell(row=i, column=1, value=a); cb = ins.cell(row=i, column=2, value=b)
        ca.font = Font(name=FONT, bold=True, size=10); cb.font = Font(name=FONT, size=10)
        ca.alignment = cb.alignment = Alignment(vertical='top', wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def importar(archivo, empresa, sucursal, usuario):
    """Carga el inventario de apertura. Cada fila = una entrada de stock.
    Devuelve {'cargadas': n_filas, 'piezas': total, 'errores': [...]}"""
    wb = load_workbook(archivo, data_only=True)
    ws = wb['Existencias'] if 'Existencias' in wb.sheetnames else wb.active

    headers = {}
    for j, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = j
    label_to_key = {label.lower(): key for label, key, _ in COLUMNAS}
    col = {label_to_key[k]: j for k, j in headers.items() if k in label_to_key}
    if 'sku' not in col or 'almacen' not in col or 'cantidad' not in col:
        raise ValueError("La plantilla debe tener al menos 'SKU', 'Almacén (código)' y 'Cantidad'.")

    productos = {p.sku: p for p in Producto.objects.filter(empresa=empresa)}
    almacenes = {}   # codigo.upper() -> Almacen
    ubicaciones = {}  # (almacen_id, codigo.upper()) -> Ubicacion

    def _almacen(codigo):
        cu = codigo.upper()
        if cu not in almacenes:
            alm, _ = Almacen.objects.get_or_create(
                sucursal=sucursal, codigo=codigo[:10],
                defaults={'empresa': empresa, 'nombre': codigo[:100]})
            almacenes[cu] = alm
        return almacenes[cu]

    def _ubicacion(almacen, codigo):
        cod = (codigo or 'GEN')[:20]
        key = (almacen.id, cod.upper())
        if key not in ubicaciones:
            ubi, _ = Ubicacion.objects.get_or_create(
                almacen=almacen, codigo=cod,
                defaults={'descripcion': cod})
            ubicaciones[key] = ubi
        return ubicaciones[key]

    errores = []
    cargadas = 0
    piezas = decimal.Decimal('0')

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def val(key):
            j = col.get(key)
            v = row[j - 1] if j and j - 1 < len(row) else None
            return str(v).strip() if v is not None else ''

        sku = val('sku')
        if not sku and not val('almacen') and not val('cantidad'):
            continue  # fila vacía
        prod = productos.get(sku)
        if not prod:
            errores.append(f"Fila {i}: SKU '{sku}' no existe en el catálogo.")
            continue
        cant = _dec(val('cantidad'))
        if cant is None or cant <= 0:
            errores.append(f"Fila {i} ({sku}): cantidad inválida.")
            continue

        serie_txt = val('serie')
        lote_txt = val('lote')
        if prod.es_serializable and not serie_txt:
            errores.append(f"Fila {i} ({sku}): el producto es serializado y falta la Serie.")
            continue
        if prod.es_serializable and cant != 1:
            errores.append(f"Fila {i} ({sku}): en serializados la Cantidad debe ser 1 (una fila por serie).")
            continue
        if prod.es_loteable and not lote_txt:
            errores.append(f"Fila {i} ({sku}): el producto es loteable y falta el Lote.")
            continue

        try:
            almacen = _almacen(val('almacen'))
            ubic = _ubicacion(almacen, val('ubicacion'))

            lote = None
            if lote_txt:
                lote, _ = Lote.objects.get_or_create(
                    producto=prod, numero_lote=lote_txt,
                    defaults={'fecha_caducidad': _fecha(val('caducidad'))})
                fc = _fecha(val('caducidad'))
                if fc and not lote.fecha_caducidad:
                    lote.fecha_caducidad = fc
                    lote.save(update_fields=['fecha_caducidad'])

            serie = None
            if serie_txt:
                serie, creada = NumeroSerie.objects.get_or_create(
                    producto=prod, serie=serie_txt, defaults={'estado': 'DISPONIBLE'})
                if not creada and serie.estado != 'DISPONIBLE':
                    errores.append(f"Fila {i} ({sku}): la serie '{serie_txt}' ya existe y no está disponible.")
                    continue

            costo = _dec(val('costo'))
            services.registrar_movimiento(
                empresa=empresa, sucursal=sucursal, producto=prod, ubicacion=ubic,
                tipo='AJUSTE_POS', origen='AJUSTE', cantidad=cant, usuario=usuario,
                lote=lote, serie=serie, referencia='CARGA INICIAL',
                costo_unitario=costo, notas='Inventario de apertura')
            # Backfill del costo del producto si venía en 0 y aquí sí lo dieron.
            if costo and costo > 0 and (prod.costo_unitario or 0) == 0:
                prod.costo_unitario = costo
                prod.save(update_fields=['costo_unitario'])
                productos[sku] = prod
            cargadas += 1
            piezas += cant
        except Exception as e:
            errores.append(f"Fila {i} ({sku}): {e}")
            continue

    return {'cargadas': cargadas, 'piezas': piezas, 'errores': errores}
