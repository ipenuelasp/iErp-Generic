"""
Plantilla y carga masiva de clientes desde Excel (.xlsx).
Mismo patrón que admon_compras/import_proveedores.py.
"""
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from admon_empresas.models import Moneda
from .models import Cliente


# (etiqueta de columna, atributo del modelo, obligatorio)
COLUMNAS = [
    ('Nombre comercial', 'nombre_comercial', False),
    ('Razón social (nombre fiscal)', 'nombre_fiscal', True),
    ('RFC', 'rfc', False),
    ('Email', 'email', False),
    ('Teléfono', 'telefono', False),
    ('Dirección', 'direccion', False),
    ('Atención con', 'contacto_nombre', False),
    ('Días crédito', 'dias_credito', False),
    ('Límite crédito', 'limite_credito', False),
    ('Moneda', 'moneda', False),
    ('Activo (Sí/No)', 'activo', False),
]

EJEMPLO = {
    'nombre_comercial': 'Hospital General',
    'nombre_fiscal': 'Hospital General, S.A. de C.V.',
    'rfc': 'HGE950101AAA',
    'email': 'compras@hospitalgeneral.mx',
    'telefono': '5551234567',
    'direccion': 'Av. Reforma 100, Cuauhtémoc, CDMX, C.P. 06600',
    'contacto_nombre': 'Dra. Ana López',
    'dias_credito': 30,
    'limite_credito': 100000,
    'moneda': 'MXN',
    'activo': 'Sí',
}

_WIDTHS = [26, 34, 16, 26, 16, 46, 20, 12, 14, 10, 13]


def generar_plantilla(con_ejemplo=True):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Clientes'
    FONT = 'Arial'
    hdr = PatternFill('solid', start_color='1F3864')
    hdr_req = PatternFill('solid', start_color='C55A11')
    thin = Side(style='thin', color='D0D0D0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for j, (label, field, req) in enumerate(COLUMNAS, start=1):
        c = ws.cell(row=1, column=j, value=label)
        c.font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
        c.fill = hdr_req if req else hdr
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        ws.column_dimensions[c.column_letter].width = _WIDTHS[j - 1]

    if con_ejemplo:
        for j, (label, field, req) in enumerate(COLUMNAS, start=1):
            c = ws.cell(row=2, column=j, value=EJEMPLO.get(field, ''))
            c.font = Font(name=FONT, size=10, italic=True, color='808080')
            c.alignment = Alignment(vertical='top', wrap_text=(field == 'direccion'))

    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30

    ins = wb.create_sheet('Instrucciones')
    ins.column_dimensions['A'].width = 30
    ins.column_dimensions['B'].width = 72
    ins.cell(row=1, column=1, value='Cómo llenar esta plantilla').font = Font(name=FONT, bold=True, size=13, color='1F3864')
    notas = [
        ('Razón social', 'OBLIGATORIO. Si no la tienes, repite el nombre comercial.'),
        ('RFC', 'Opcional. Si se repite uno existente, se actualiza ese cliente.'),
        ('Días crédito', 'Número entero. 0 = de contado.'),
        ('Límite crédito', 'Número. 0 = sin límite.'),
        ('Moneda', 'MXN o USD. Vacío = MXN.'),
        ('Activo (Sí/No)', 'Sí = alta. Vacío también es Sí.'),
        ('Fila de ejemplo', 'Bórrala antes de cargar.'),
    ]
    for i, (a, b) in enumerate(notas, start=3):
        ca = ins.cell(row=i, column=1, value=a); cb = ins.cell(row=i, column=2, value=b)
        ca.font = Font(name=FONT, bold=True, size=10); cb.font = Font(name=FONT, size=10)
        ca.alignment = cb.alignment = Alignment(vertical='top', wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _norm_bool(v):
    if v is None or str(v).strip() == '':
        return True
    return str(v).strip().lower() in ('sí', 'si', 's', 'x', '1', 'true', 'verdadero', 'activo')


def _num(v, entero=False):
    s = str(v).strip() if v is not None else ''
    if not s:
        return 0
    try:
        return int(float(s)) if entero else float(s)
    except ValueError:
        return 0


def importar(archivo, empresa):
    wb = load_workbook(archivo, data_only=True)
    ws = wb['Clientes'] if 'Clientes' in wb.sheetnames else wb.active

    headers = {}
    for j, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = j
    label_to_field = {label.lower(): field for label, field, _ in COLUMNAS}
    col = {label_to_field[k]: j for k, j in headers.items() if k in label_to_field}
    if 'nombre_fiscal' not in col:
        raise ValueError("La plantilla no tiene la columna 'Razón social (nombre fiscal)'.")

    monedas = {m.codigo.upper(): m for m in Moneda.objects.filter(empresa=empresa)}
    creados = actualizados = 0
    errores = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        def val(field):
            j = col.get(field)
            v = row[j - 1] if j and j - 1 < len(row) else None
            return str(v).strip() if v is not None else ''

        nombre_fiscal = val('nombre_fiscal')
        nombre_comercial = val('nombre_comercial')
        if not nombre_fiscal and not nombre_comercial:
            continue
        if not nombre_fiscal:
            nombre_fiscal = nombre_comercial
        rfc = val('rfc')

        datos = {
            'nombre_comercial': nombre_comercial or None,
            'rfc': rfc,
            'email': val('email') or None,
            'telefono': val('telefono'),
            'direccion': val('direccion'),
            'contacto_nombre': val('contacto_nombre'),
            'dias_credito': _num(val('dias_credito'), entero=True),
            'limite_credito': _num(val('limite_credito')),
            'activo': _norm_bool(val('activo')),
            'moneda_predeterminada': monedas.get(val('moneda').upper() or 'MXN'),
        }
        try:
            existente = None
            if rfc:
                existente = Cliente.objects.filter(empresa=empresa, rfc=rfc).first()
            if not existente:
                existente = Cliente.objects.filter(empresa=empresa, nombre_fiscal=nombre_fiscal).first()
            if existente:
                for k, v in datos.items():
                    setattr(existente, k, v)
                existente.save()
                actualizados += 1
            else:
                Cliente.objects.create(empresa=empresa, nombre_fiscal=nombre_fiscal, **datos)
                creados += 1
        except Exception as e:
            errores.append(f"Fila {i} ({nombre_fiscal}): {e}")

    return {'creados': creados, 'actualizados': actualizados, 'errores': errores}
